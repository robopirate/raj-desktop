# --- Calendar & Drive Integration (v4.0) ---
CALENDAR_AVAILABLE = False
DRIVE_AVAILABLE = False
try:
    from calendar_integration import CalendarManager
    CALENDAR_AVAILABLE = True
except ImportError:
    pass
try:
    from drive_integration import DriveManager
    DRIVE_AVAILABLE = True
except ImportError:
    pass
# --- End Integration ---

"""
engine.py -- RoboPirate Campaign Engine v5.0
SCHOOL + CSR sequences | Raj as manager | Auto-send | Draft-only replies
FIXED: HTML template, auto-advance scheduling, template sync regex
"""

import re
import json
import time
import threading
import email.utils
import html as html_module
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
from collections import deque

from db import Database
from gmail import GmailClient
from tracking_server import TrackingServer
from notifications import notify

try:
    from smart_importer import SmartImporter
    SMART_IMPORT_AVAILABLE = True
except ImportError:
    SMART_IMPORT_AVAILABLE = False

try:
    from rewritten_email_templates import (
        _generate_school_content as _new_school_content,
        _generate_school_text_content as _new_school_text_content,
        _generate_csr_wsl5_content as _new_csr_wsl5_content,
        _generate_csr_wsl5_text_content as _new_csr_wsl5_text_content,
        REWRITTEN_SUBJECTS,
        PREHEADERS,
    )
    REWRITTEN_TEMPLATES_AVAILABLE = True
except ImportError:
    REWRITTEN_TEMPLATES_AVAILABLE = False

# Sequences: SCHOOL (private schools) and CSR (corporates)
SEQUENCES = {
    "school": {
        "days": [1, 3, 5, 7, 10],
        "template_prefix": "SCHOOL EMAIL ",
        "audience": "private_school",
        "persona": "school",
        "assets": {
            1: {
                "brochure": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_wsl": "https://www.instagram.com/p/DTDBcsdk9FI/",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_ig": "https://www.instagram.com/robo.pirate/"
            },
            3: {
                "report_vbv": "https://drive.google.com/file/d/1d7EEtC8YitbSj7U6ivHf_6WtUGuylT-B/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA"
            },
            5: {
                "report_1st_wsl": "https://drive.google.com/file/d/1H7mHVTWGprbd4ZFSPoJZPeAc1nHnih3J/view"
            },
            7: {
                "report_sangli1": "https://drive.google.com/file/d/1HpNdnamA2k3H0xkKr58STEKMNu5RgHPx/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA?si=ZFAr_bp_xU2Sduwr",
                "video_star": "https://www.youtube.com/watch?v=iziKPBSfGKU",
                "video_bandhuta": "https://www.youtube.com/watch?v=xVmaBnPyg9A",
                "video_sbn": "https://www.youtube.com/watch?v=d-TsgUkhIu0",
                "video_we": "https://www.instagram.com/reel/DMe2HzqofAk/?igsh=c201ZGxsOGFlMjJj"
            },
            10: {
                "plans": "https://drive.google.com/file/d/1p2CyHVZK_giZj0KNDGTTs_-s7HxVnQ_C/view"
            }
        }
    },
    "csr": {
        "days": [1, 3, 5, 7, 10],
        "template_prefix": "CSR EMAIL ",
        "audience": "csr",
        "persona": "csr",
        "assets": {
            1: {
                "report_sangli1": "https://drive.google.com/file/d/1HpNdnamA2k3H0xkKr58STEKMNu5RgHPx/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_sangli": "https://drive.google.com/file/d/1MUlsC87vRbhFaoW0XcX146WBLKYBk448/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            3: {
                "report_sangli1": "https://drive.google.com/file/d/1HpNdnamA2k3H0xkKr58STEKMNu5RgHPx/view",
                "brochure": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            5: {
                "report_sangli2": "https://drive.google.com/file/d/1pKSm1WPlPk-we4aC-uhqxEy8w-BYygSN/view",
                "report_vbv": "https://drive.google.com/file/d/1d7EEtC8YitbSj7U6ivHf_6WtUGuylT-B/view",
                "video_star": "https://youtube.com/watch?v=iziKPBSfGKU",
                "folder_sangli": "https://drive.google.com/drive/folders/15sc5iOIKTBZyenb2rCpGVAK1lExcG5BC",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            7: {
                "plans": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view",
                "video_wsl": "https://drive.google.com/file/d/1KPrC2IpdooxazGJiyVe79JgyWlJbOxzu/view",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_sangli": "https://drive.google.com/file/d/1MUlsC87vRbhFaoW0XcX146WBLKYBk448/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            10: {
                "profile": "https://drive.google.com/file/d/1g9JJ4_VO_28QKYD7iVVDJZcv9l4uRbZu/view",
                "kits": "https://drive.google.com/file/d/1cvi4p8IHgx1MekanVRHN3Fo4Lk9vbubX/view",
                "video_ig": "https://www.instagram.com/reel/DMe2HzqofAk/"
            }
        }
    },
    "csr-wsl-5": {
        "days": [1, 3, 5, 7, 10],
        "template_prefix": "CSR-WSL-5 EMAIL ",
        "audience": "csr",
        "persona": "csr",
        "assets": {
            1: {
                "brochure": "https://drive.google.com/file/d/1vRMeFM22aajc5zfiYhqaev34UVQ87zyU/view"
            },
            3: {
                "video_wsl": "https://www.instagram.com/p/DTDBcsdk9FI/"
            },
            5: {
                "report_1st_wsl": "https://drive.google.com/file/d/1H7mHVTWGprbd4ZFSPoJZPeAc1nHnih3J/view",
                "report_vbv": "https://drive.google.com/file/d/1d7EEtC8YitbSj7U6ivHf_6WtUGuylT-B/view"
            },
            7: {
                "report_sangli": "https://drive.google.com/file/d/1pKSm1WPlPk-we4aC-uhqxEy8w-BYygSN/view",
                "video_divyang": "https://www.instagram.com/p/DMhEDutOrl-/",
                "video_gruh": "https://www.instagram.com/p/DSSIy7nglXc/",
                "video_abp": "https://youtu.be/FJ2_W53WjmA",
                "video_star": "https://www.youtube.com/watch?v=iziKPBSfGKU",
                "video_bandhuta": "https://www.youtube.com/watch?v=xVmaBnPyg9A",
                "video_sbn": "https://www.youtube.com/watch?v=d-TsgUkhIu0",
                "video_we": "https://www.instagram.com/reel/DMe2HzqofAk/"
            },
            10: {
                "proposal_2nd": "https://drive.google.com/file/d/1NdMn4J8DgWyoNMyTHUv2t2caP3wTYIkq/view"
            }
        }
    }
}

EMAIL_NUM_TO_DAY = {1: 1, 2: 3, 3: 5, 4: 7, 5: 10}
DAY_TO_EMAIL_NUM = {1: 1, 3: 2, 5: 3, 7: 4, 10: 5}
VERSION = "5.0.0"

SEND_DELAY = 45
BOUNCE_INTERVAL = 6
REPLY_INTERVAL = 60
EMERGENCY_INTERVAL = 15
EOD_HOUR = 19
MORNING_HOUR = 8

HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Robo Pirate - WE Smart Lab</title>
</head>
<body style="margin:0;padding:0;background-color:#F5F9F9;font-family:Arial,Helvetica,sans-serif;color:#333333;">
<div style="display:none;font-size:1px;line-height:1px;max-height:0;max-width:0;opacity:0;overflow:hidden;mso-hide:all;">{preheader}</div>
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#F5F9F9">
<tr>
<td align="center" style="padding:20px 10px;">
<table role="presentation" width="600" cellpadding="0" cellspacing="0" border="0" style="max-width:600px;width:100%;background-color:#FFFFFF;border-radius:8px;overflow:hidden;">
<tr>
<td align="center" bgcolor="#FF2E88" style="padding:8px 20px;">
<img src="https://drive.google.com/thumbnail?id=1krjWmmgui9h1V6MRWKOLsXP2JpBfPrFk&amp;sz=w1000" alt="WE Smart Lab" width="265" style="display:block;max-width:100%;height:auto;border:0;">
</td>
</tr>
<tr><td bgcolor="#FFD400" style="height:4px;line-height:4px;font-size:0;background:linear-gradient(90deg,#FFD400,#9333EA);">&nbsp;</td></tr>
<tr>
<td style="padding:30px 25px;font-size:15px;line-height:1.6;color:#333333;">
{body}
</td>
</tr>
<tr>
<td style="padding:15px 25px;background-color:#F5F9F9;border-top:1px solid #E0E8E8;text-align:center;font-size:12px;color:#7A8A8A;line-height:1.4;">
Robo Pirate, Baner–Mahalunge Road, Opp. Shreeram Sankul, Next to Euphoria Bungalow,<br>
Baner Gaon, Baner, Pune, Maharashtra 411045, India · Phone: +91 91368 99925 · info@robopirate.in<br>
To stop receiving these emails, reply STOP and we will remove you within 48 hours.
</td>
</tr>
</table>
</td>
</tr>
</table>
</body>
</html>"""

@dataclass
class Recipient:
    id: int
    sequence_id: str
    email: str
    name: str
    org: str
    extra_json: str
    sub_pool: str = ""
    import_status: str = ""
    import_error: str = ""
    imported_at: str = ""
    batched: int = 0

@dataclass
class BatchResult:
    queued: int
    sent: int = 0
    drafted: int = 0
    error: Optional[str] = None

class CampaignEngine:
    def __init__(self, db: Database, gmail: GmailClient, ollama_url="http://localhost:11434"):
        self.db = db
        self.gmail = gmail
        self.ollama_url = ollama_url
        self._running = False
        self._thread = None
        self._paused = False
        self._last_batch_process_time = None
        self.brief_email = db.get_meta("brief_email") or ""
        self.calendar = CalendarManager() if CALENDAR_AVAILABLE else None
        self.drive = DriveManager() if DRIVE_AVAILABLE else None
        self.logs = deque(maxlen=200)
        self._log_callbacks = []
        self.tracker = None

    def _get_verified_sender(self) -> Optional[str]:
        """Return a verified From address, or None to let Gmail use the authenticated account."""
        wanted = (self.db.get_meta("default_sender") or "").strip()
        if not wanted:
            return None
        if not self.gmail or not self.gmail.is_connected():
            # Cannot verify yet; do not risk silent rewrite by Gmail.
            return None
        try:
            send_as = self.gmail.service.users().settings().sendAs().list(userId="me").execute()
            aliases = {a.get("sendAsEmail", "").lower() for a in send_as.get("sendAs", [])}
            if wanted.lower() in aliases:
                return wanted
        except Exception as e:
            self._log(f"[Sender] Could not verify send-as aliases: {e}")
        return None

    @property
    def default_sender(self):
        return self._get_verified_sender()

    def add_log_callback(self, fn):
        self._log_callbacks.append(fn)

    def connect_gmail(self, callback=None):
        """Trigger interactive Gmail OAuth in a background thread."""
        if not self.gmail:
            if callback:
                callback(False, "Gmail client not available")
            return
        def cb(success, error):
            if success:
                self._log("[Gmail] Connected")
            else:
                self._log(f"[Gmail] Connection failed: {error}")
            if callback:
                callback(success, error)
        threading.Thread(target=self.gmail.authenticate, args=(cb,), daemon=True).start()

    def connect_calendar(self, callback=None):
        """Trigger interactive Calendar OAuth in a background thread."""
        if not self.calendar:
            if callback:
                callback(False, "Calendar integration not available")
            return
        def cb(success, error):
            if success:
                self._log("[Calendar] Connected")
            else:
                self._log(f"[Calendar] Connection failed: {error}")
            if callback:
                callback(success, error)
        threading.Thread(target=self.calendar.authenticate, args=(cb,), daemon=True).start()

    def connect_drive(self, callback=None):
        """Trigger interactive Drive OAuth in a background thread."""
        if not self.drive:
            if callback:
                callback(False, "Drive integration not available")
            return
        def cb(success, error):
            if success:
                self._log("[Drive] Connected")
            else:
                self._log(f"[Drive] Connection failed: {error}")
            if callback:
                callback(success, error)
        threading.Thread(target=self.drive.authenticate, args=(cb,), daemon=True).start()

    def _log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        self.logs.append(line)
        print(f"[Engine] {msg}")
        for fn in self._log_callbacks:
            try:
                fn(line)
            except:
                pass

    @staticmethod
    def html_to_text(html_body: str) -> str:
        """Convert HTML email body to clean plain text for multipart sending.
        Preserves structure, links, and readability. Idempotent on plain text."""
        if not html_body:
            return ""
        # If already plain text (no real HTML tags), return as-is
        if not re.search(r'<[a-zA-Z/][^>]*>', html_body):
            return html_body.strip()

        text = html_body
        # Strip invisible/head blocks first so CSS/JS does not leak into plain text
        text = re.sub(r'<(style|script)[^>]*>.*?</\1>', '', text, flags=re.IGNORECASE | re.DOTALL)
        # Replace <br> variants with newlines
        text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
        # Replace </p> with double newline
        text = re.sub(r'</p>', '\n\n', text, flags=re.IGNORECASE)
        # Replace </div> with newline
        text = re.sub(r'</div>', '\n', text, flags=re.IGNORECASE)
        # Replace </li> with newline + bullet marker
        text = re.sub(r'</li>', '\n', text, flags=re.IGNORECASE)
        # Replace <li> with bullet
        text = re.sub(r'<li[^>]*>', '• ', text, flags=re.IGNORECASE)
        # Replace <h1>-<h6> with uppercase + newlines
        for i in range(6, 0, -1):
            text = re.sub(rf'<h{i}[^>]*>(.*?)</h{i}>', lambda m: f'\n\n{m.group(1).upper()}\n{"=" * len(m.group(1))}\n', text, flags=re.IGNORECASE | re.DOTALL)
        # Replace <strong>, <b> with **text**
        text = re.sub(r'<(strong|b)[^>]*>(.*?)</\1>', r'**\2**', text, flags=re.IGNORECASE | re.DOTALL)
        # Replace <em>, <i> with _text_
        text = re.sub(r'<(em|i)[^>]*>(.*?)</\1>', r'_\2_', text, flags=re.IGNORECASE | re.DOTALL)
        # Replace <a href="...">text</a> with text (URL)
        text = re.sub(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', lambda m: f'{m.group(2)} ({m.group(1)})', text, flags=re.IGNORECASE | re.DOTALL)
        # Remove remaining HTML tags
        text = re.sub(r'<[^>]+>', '', text)
        # Decode HTML entities
        text = html_module.unescape(text)
        # Clean up excessive whitespace
        text = re.sub(r'\n{3,}', '\n\n', text)
        text = re.sub(r'[ \t]+', ' ', text)
        return text.strip()

    def _notify(self, title: str, message: str, timeout: int = 5):
        """Send a desktop notification if enabled in meta settings."""
        try:
            val = self.db.get_meta("desktop_notifications")
            if val is not None and str(val).lower() in ("false", "0", "off"):
                return
            notify(title, message, timeout)
        except Exception:
            pass

    # -- Lifecycle --
    def start(self):
        if self._running: return
        self._running = True

        # Start tracking server for engagement analytics
        try:
            public_url = self.db.get_meta("tracking_public_url") or None
            self.tracker = TrackingServer(self.db.db_path, public_base_url=public_url)
            self.tracker.start()
            if self.tracker.base_url:
                self._log(f"[Tracking] Engagement tracking active at {self.tracker.base_url}")
            else:
                self._log("[Tracking] Local-only server; tracking injection disabled (set tracking_public_url meta to enable)")
        except Exception as e:
            self._log(f"[Tracking] Failed to start: {e}")

        # TEMPLATE HEALTH: ensure built-in sequences have valid templates
        try:
            health = self.validate_templates(auto_repair=True)
            if health["repaired"]:
                self._log(f"[TemplateHealth] Auto-repaired: {', '.join(health['repaired'])}")
            if health["failed"]:
                self._log(f"[TemplateHealth] Still broken: {', '.join(health['failed'])}")
            if health["ok"]:
                self._log("[TemplateHealth] All templates valid")
        except Exception as e:
            self._log(f"[TemplateHealth] Check failed: {e}")

        # RESUME-ON-BOOT: Check for batches stuck in "running" status
        try:
            running_batches = self.db.get_running_batches()
            if running_batches:
                self._log(f"[RESUME] Found {len(running_batches)} batch(es) in RUNNING status from previous session")
                for batch in running_batches:
                    self._log(f"[RESUME] Will continue batch '{batch['name']}' (ID: {batch['id']})")
            else:
                self._log("[RESUME] No running batches from previous session")

            # Also check scheduled batches that may have missed their time
            scheduled_batches = self.db.get_scheduled_batches()
            if scheduled_batches:
                now = datetime.now()
                missed = 0
                for batch in scheduled_batches:
                    sched_str = batch.get("scheduled_at")
                    if sched_str:
                        try:
                            sched_dt = datetime.fromisoformat(sched_str)
                            if now > sched_dt:
                                missed += 1
                                self._log(f"[RESUME] Batch '{batch['name']}' missed schedule ({sched_dt.strftime('%d %b %H:%M')}) — will auto-start")
                        except:
                            pass
                if missed > 0:
                    self._log(f"[RESUME] {missed} scheduled batch(es) missed their time while system was off")
        except Exception as e:
            self._log(f"[RESUME] Error checking previous state: {e}")

        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()
        self._log("Raj Engine started")

    def stop(self):
        self._running = False
        self._log("Engine stopping...")

    def pause(self):
        self._paused = True
        self._log("PAUSED")

    def resume(self):
        self._paused = False
        self._log("RESUMED")

    def is_running(self): return self._running
    def is_paused(self): return self._paused

    # -- Main Loop --
    def _loop(self):
        while self._running:
            try:
                if not self._paused:
                    self._tick()
            except Exception as e:
                self._log(f"LOOP ERROR: {e}")
            time.sleep(60)

    def _tick(self):
        now = datetime.now()
        self._check_scheduled_sends(now)
        self._process_running_batches(now)
        self._check_auto_start_scheduled_batches(now)
        self._check_bounce_scan(now)
        self._check_reply_scan(now)
        self._check_emergency_commands(now)
        self._check_eod(now)
        self._check_morning_brief(now)

    # -- Batch Processing (NEW) --
    def _process_running_batches(self, now: datetime):
        """Process running batches and send emails at staggered intervals."""
        try:
            running_batches = self.db.execute(
                "SELECT * FROM batches WHERE status='running' AND deleted_at IS NULL ORDER BY created_at"
            ).fetchall()

            if not running_batches:
                return

            # BATCH SLOT LOCK: Only process one batch at a time
            if self._last_batch_process_time:
                if (now - self._last_batch_process_time).total_seconds() < 2:
                    return
            self._last_batch_process_time = now

            for batch_row in running_batches:
                batch = dict(batch_row)
                batch_id = batch["id"]
                seq_id = batch["sequence_id"]
                if seq_id and self.db.get_meta(f"pause_{seq_id}") == "true":
                    self._log(f"[Batch {batch_id}] {seq_id.upper()} is paused, skipping")
                    continue
                if seq_id == "unassigned":
                    self._log(f"[Batch {batch_id}] UNASSIGNED — pausing. Assign a sequence before running.")
                    self.db.batch_update_status(batch_id, "paused")
                    continue
                stagger = batch.get("stagger_minutes", 0) or 1
                day_offset = batch.get("day_offset", 1)

                # Find next pending recipient
                next_recipient = self.db.execute("""
                    SELECT r.id, r.sequence_id, r.email, r.name, r.org, r.extra_json, r.sub_pool, r.import_status, r.import_error, r.imported_at, r.batched
                    FROM recipients r
                    JOIN batch_recipients br ON r.id = br.recipient_id
                    WHERE br.batch_id = ? AND br.status = 'pending'
                    ORDER BY r.id
                    LIMIT 1
                """, (batch_id,)).fetchone()

                if not next_recipient:
                    # All sent - mark completed and auto-advance
                    self.db.batch_update_status(batch_id, "completed")
                    counts = self.db.batch_count_by_status(batch_id)
                    total = sum(counts.values())
                    sent = counts.get("sent", 0)
                    self._log(f"[Batch {batch_id}] Completed: all recipients processed")
                    self._notify("Batch Complete", f"'{batch.get('name', batch_id)}' done. {sent}/{total} sent.")
                    self._auto_advance_batch(batch)
                    continue

                # Check if enough time passed since last send
                last_send = self.db.execute("""
                    SELECT MAX(sent_at) FROM batch_recipients
                    WHERE batch_id = ? AND status = 'sent'
                """, (batch_id,)).fetchone()[0]

                if last_send:
                    last_dt = datetime.fromisoformat(last_send)
                    minutes_since = (now - last_dt).total_seconds() / 60
                    if minutes_since < stagger:
                        continue

                # BLACKLIST CHECK: Skip if email was blacklisted
                rec_email = next_recipient[2]
                if self.db.blacklist_has(rec_email):
                    self._log(f"[Batch {batch_id}] SKIPPING blacklisted: {rec_email}")
                    self.db.execute("""
                        UPDATE batch_recipients SET status='skipped'
                        WHERE batch_id=? AND recipient_id=?
                    """, (batch_id, next_recipient[0]))
                    self.db.commit()
                    continue

                # SUNDAY FILTER: Skip sends on Sunday
                if now.weekday() == 6:
                    self._log(f"[Batch {batch_id}] SUNDAY — skipping send for {rec_email}, will resume Monday")
                    continue

                # DEDUP: Skip if this recipient already received this sequence/day
                already_sent = self.db.execute("""
                    SELECT 1 FROM sends s
                    JOIN recipients r ON s.recipient_id = r.id
                    WHERE r.email = ? AND s.day = ? AND s.status = 'sent'
                    LIMIT 1
                """, (rec_email.lower().strip(), day_offset)).fetchone()
                if already_sent:
                    self._log(f"[DEDUP] {rec_email} already sent Day {day_offset} — marking skipped in batch {batch_id}")
                    self.db.execute("""
                        UPDATE batch_recipients SET status='sent'
                        WHERE batch_id=? AND recipient_id=?
                    """, (batch_id, next_recipient[0]))
                    self.db.commit()
                    continue

                rec = Recipient(*next_recipient)

                # REPLY CHECK: Stop sequence if recipient has already replied
                replied = self.db.execute(
                    "SELECT 1 FROM replies WHERE recipient_id=? AND status IN ('pending','drafted','handled') LIMIT 1",
                    (rec.id,)
                ).fetchone()
                if replied:
                    self.stop_sequence_for_recipient(rec.id, f"{rec.email} already replied")
                    self.db.execute("""
                        UPDATE batch_recipients SET status='replied'
                        WHERE batch_id=? AND recipient_id=?
                    """, (batch_id, rec.id))
                    self.db.commit()
                    self._log(f"[SKIP] {rec.email} already replied — stopping sequence")
                    continue

                # BOUNCE CHECK: Stop sequence if the latest send for this day bounced
                bounced = self.db.execute(
                    "SELECT status FROM sends WHERE recipient_id=? AND day=? ORDER BY id DESC LIMIT 1",
                    (rec.id, day_offset)
                ).fetchone()
                if bounced and bounced[0] == 'bounced':
                    self.stop_sequence_for_recipient(rec.id, f"{rec.email} bounced")
                    self._log(f"[SKIP] {rec.email} bounced — stopping sequence")
                    continue

                # Send email
                subj, body_html, body_text, ab_variant, fmt = self.render(seq_id, day_offset, rec)
                if not subj:
                    self._log(f"[Batch {batch_id}] No template for {rec.email} Day {day_offset}, skipping")
                    self.db.execute("UPDATE batch_recipients SET status='failed' WHERE batch_id=? AND recipient_id=?",
                        (batch_id, rec.id))
                    self.db.commit()
                    continue

                # Empty body guard: do not send blank emails
                is_plain = fmt == 'plain'
                if (is_plain and not (body_text or "").strip()) or (not is_plain and not (body_html or "").strip()):
                    self._log(f"[Batch {batch_id}] Empty body for {rec.email} Day {day_offset}, skipping")
                    self.db.execute("UPDATE batch_recipients SET status='failed' WHERE batch_id=? AND recipient_id=?",
                        (batch_id, rec.id))
                    self.db.commit()
                    continue

                try:
                    # Determine if draft or immediate send
                    sched_str = batch.get("scheduled_at")
                    use_draft = False
                    if sched_str:
                        try:
                            sched_dt = datetime.fromisoformat(sched_str.replace("Z", "+00:00"))
                            use_draft = sched_dt > now
                        except:
                            pass

                    # Pre-insert sends record to get send_id for tracking
                    placeholder_status = "drafted" if use_draft else "pending"
                    send_id = self.db.campaign_queue_send(rec.id, day_offset, subj, "pending", placeholder_status, batch_id, ab_variant)

                    # Inject tracking pixel and wrapped links with real send_id (HTML only)
                    if self.tracker and self.tracker.base_url and send_id and fmt != 'plain':
                        body_html = self.tracker.inject_tracking(body_html, rec.id, batch_id, send_id)

                    if use_draft:
                        draft = self.gmail.create_scheduled_draft(rec.email, subj, body_html, sched_str, body_text, sender=self.default_sender, format=fmt)
                        self.db.execute("""
                            UPDATE batch_recipients SET status='drafted', sent_at=?
                            WHERE batch_id=? AND recipient_id=?
                        """, (now.isoformat(), batch_id, rec.id))
                        self.db.execute("UPDATE sends SET draft_id=?, status='drafted', ab_variant=? WHERE id=?",
                                        (draft.get("id"), ab_variant, send_id))
                        self.db.commit()
                        self._log(f"[Batch {batch['name']}] Scheduled draft for {rec.email} ({seq_id.upper()} Day {day_offset})")
                    else:
                        # Delete any old scheduled draft before sending for real
                        old_draft = self.db.execute("""
                            SELECT draft_id FROM sends 
                            WHERE recipient_id=? AND batch_id=? AND status='drafted' AND draft_id IS NOT NULL
                            ORDER BY id DESC LIMIT 1
                        """, (rec.id, batch_id)).fetchone()
                        if old_draft and old_draft[0]:
                            try:
                                self.gmail.delete_draft(old_draft[0])
                                self._log(f"[Batch {batch['name']}] Deleted old draft for {rec.email}")
                            except Exception as del_err:
                                self._log(f"[Batch {batch['name']}] Draft delete skipped: {del_err}")

                        msg = self._send_with_retry(rec.email, subj, body_html, body_text, format=fmt)
                        self.db.execute("""
                            UPDATE batch_recipients SET status='sent', sent_at=?
                            WHERE batch_id=? AND recipient_id=?
                        """, (now.isoformat(), batch_id, rec.id))
                        self.db.execute("UPDATE sends SET draft_id=?, status='sent', sent_at=?, ab_variant=? WHERE id=?",
                                        (msg.get("id"), now.isoformat(), ab_variant, send_id))
                        self.db.commit()
                        self._log(f"[Batch {batch['name']}] Sent to {rec.email} ({seq_id.upper()} Day {day_offset})")
                except Exception as e:
                    err = str(e).lower()
                    self._log(f"[Batch {batch_id}] Failed to send to {rec.email}: {e}")
                    self.db.execute("UPDATE batch_recipients SET status='failed' WHERE batch_id=? AND recipient_id=?",
                        (batch_id, rec.id))
                    self.db.execute("UPDATE sends SET status='failed' WHERE id=?", (send_id,))
                    self.db.commit()
                    # Only blacklist on Gmail 4xx recipient errors (e.g. invalidRecipient).
                    # Auth/network errors must NOT blacklist valid leads.
                    is_recipient_error = False
                    status = getattr(e, 'status', None)
                    reason = getattr(e, 'reason', '') or ''
                    if status in (400, 403, 404) and any(k in (reason.lower() + err) for k in ['invalidrecipient', 'badaddress', 'recipient', 'address not found']):
                        is_recipient_error = True
                    if is_recipient_error:
                        self.db.blacklist_add(rec.email, f"send_failed:{err[:80]}")
                        self._log(f"[BLACKLIST] {rec.email} — permanent failure")

        except Exception as e:
            self._log(f"DEBUG ERROR in _process_running_batches: {e}")
            import traceback
            self._log(f"DEBUG TRACEBACK: {traceback.format_exc()}")

    def stop_sequence_for_recipient(self, recipient_id, reason):
        """Stop all pending sends for a recipient across all batches in their family."""
        rows = self.db.execute("""
            SELECT b.id, b.name
            FROM batches b
            JOIN batch_recipients br ON b.id = br.batch_id
            WHERE br.recipient_id = ? AND br.status = 'pending'
        """, (recipient_id,)).fetchall()

        updated = 0
        for batch_row in rows:
            batch_id = batch_row["id"]
            cur = self.db.execute("""
                UPDATE batch_recipients SET status='stopped'
                WHERE batch_id=? AND recipient_id=? AND status='pending'
            """, (batch_id, recipient_id))
            updated += cur.rowcount

        self.db.commit()
        if updated:
            self._log(f"[STOP] recipient {recipient_id}: {reason} ({updated} pending days stopped)")
        return updated

    def _auto_advance_batch(self, completed_batch: dict):
        """Auto-create next day batch and schedule it. Like Brevo — next day starts after delay."""
        seq_id = completed_batch["sequence_id"]
        current_day = completed_batch.get("day_offset", 1)
        cfg = SEQUENCES.get(seq_id)
        if not cfg:
            return

        days = cfg["days"]
        try:
            idx = days.index(current_day)
        except ValueError:
            return

        if idx >= len(days) - 1:
            self._log(f"[AUTO-ADVANCE] {seq_id.upper()} sequence complete! All {len(days)} days done.")
            return

        next_day = days[idx + 1]
        parent_name = completed_batch["name"]
        base_name = parent_name.split("-D")[0] if "-D" in parent_name else parent_name
        next_name = f"{base_name}-D{next_day}"

        # DEDUP: Skip if next-day batch already exists
        existing = self.db.execute(
            "SELECT id FROM batches WHERE name=? AND day_offset=? AND sequence_id=?",
            (next_name, next_day, seq_id)
        ).fetchone()
        if existing:
            self._log(f"[AUTO-ADVANCE] Skip: {next_name} already exists (ID: {existing[0]})")
            return

        # Schedule for +2 days at 10 AM (from completion time, not now)
        completed_at = completed_batch.get('completed_at')
        if completed_at:
            try:
                base_dt = datetime.fromisoformat(completed_at)
            except:
                base_dt = datetime.now()
        else:
            base_dt = datetime.now()
        scheduled = (base_dt + timedelta(days=2)).replace(hour=10, minute=0, second=0, microsecond=0)

        # Get parent_batch_id (link to original batch)
        parent_batch_id = completed_batch.get("parent_batch_id") or completed_batch["id"]

        # Get recipients from the completed batch to copy to next batch
        prev_recipients = self.db.batch_get_recipients(completed_batch["id"])

        # FIX: Create batch with status='scheduled' so auto-start picks it up
        new_batch_id = self.db.batch_create(
            next_name, seq_id, scheduled.isoformat(),
            stagger_minutes=completed_batch.get("stagger_minutes", 2),
            day_offset=next_day,
            parent_batch_id=parent_batch_id
        )

        # FIX: Set status to 'scheduled' after creation (batch_create defaults to 'draft')
        self.db.execute("UPDATE batches SET status='scheduled' WHERE id=?", (new_batch_id,))
        self.db.commit()

        # Only carry forward recipients who were successfully sent AND not blacklisted
        carried = 0
        for r in prev_recipients:
            if r.get("batch_status") == "sent":
                # Skip if email was blacklisted after send (bounce, etc.)
                if self.db.blacklist_has(r.get("email", "")):
                    continue
                self.db.batch_add_recipient(new_batch_id, r["id"])
                carried += 1

        self._log(f"[AUTO-ADVANCE] Created {next_name} for {scheduled.strftime('%d %b %H:%M')} ({carried}/{len(prev_recipients)} recipients carried forward)")
        self._log(f"[AUTO-ADVANCE] Pipeline: {base_name} Day {current_day} → Day {next_day} (parent: {parent_batch_id})")

    def _check_auto_start_scheduled_batches(self, now: datetime):
        """Auto-start scheduled batches when their time arrives.
        Includes draft/paused batches that have a scheduled_at date set."""
        scheduled = self.db.execute("""
            SELECT * FROM batches
            WHERE scheduled_at IS NOT NULL AND scheduled_at != ''
              AND status IN ('scheduled', 'draft', 'paused')
              AND deleted_at IS NULL
        """).fetchall()

        for batch_row in scheduled:
            batch = dict(batch_row)
            sched_str = batch.get("scheduled_at")
            if not sched_str:
                continue
            try:
                sched_dt = datetime.fromisoformat(sched_str)
                if now >= sched_dt:
                    # DEDUP: Don't start if another batch with same name/day/seq is already running
                    dup = self.db.execute("""
                        SELECT id FROM batches
                        WHERE name=? AND day_offset=? AND sequence_id=? AND status='running' AND id != ?
                        LIMIT 1
                    """, (batch["name"], batch.get("day_offset", 1), batch["sequence_id"], batch["id"])).fetchone()
                    if dup:
                        self.db.batch_update_status(batch["id"], "paused")
                        self._log(f"[AUTO-START] Skip '{batch['name']}' — duplicate already running (ID: {dup[0]})")
                        continue
                    self.db.batch_update_status(batch["id"], "running")
                    self._log(f"[AUTO-START] Batch '{batch['name']}' is now running")
            except:
                pass

    # -- Scheduled Sends (10 AM, auto-send sequences) --
    def _check_scheduled_sends(self, now: datetime):
        if now.weekday() == 6:
            return

        today_10am = now.replace(hour=10, minute=0, second=0, microsecond=0)
        if now < today_10am:
            return
        last = self.db.get_meta("last_scheduled_send_date")
        if last == now.strftime("%Y-%m-%d"):
            return

        self._log(f"Scheduled send check: {today}")
        for seq_id in SEQUENCES:
            if self.db.get_meta(f"pause_{seq_id}") == "true":
                self._log(f"{seq_id.upper()} is paused, skipping")
                continue
            for day in SEQUENCES[seq_id]["days"]:
                due = self.due_recipients(seq_id, day)
                if due:
                    self._log(f"{seq_id.upper()} Day {day}: {len(due)} due. Auto-sending...")
                    result = self.send_batch(seq_id, day)
                    self._log(f"Sent {result.sent}/{result.queued}")
        self.db.set_meta("last_scheduled_send_date", today)

    # -- Import --
    def smart_import(self, filepath: str, sequence_id: str = None, sub_pool: str = None) -> dict:
        """Smart import to POOL only (no batch creation). Leads go to DB first."""
        if not SMART_IMPORT_AVAILABLE:
            return {"success": False, "error": "smart_importer.py not available"}
        try:
            importer = SmartImporter(self.db, self)
            return importer.import_to_pool(filepath, sequence_id or "leads", sub_pool=sub_pool)
        except Exception as e:
            self._log(f"Smart import error: {e}")
            return {"success": False, "error": str(e)}

    def import_recipients(self, path: str, sequence_id: str = None, mapping: dict = None, sub_pool: str = None) -> Tuple[int, int]:
        mapping = mapping or {}
        try:
            import openpyxl
        except ImportError:
            self._log("openpyxl not installed. Run: pip install openpyxl")
            return 0, 0

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        imported, skipped = 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            email = str(row_dict.get(mapping.get("email", "Email"), "")).strip().lower()
            name = str(row_dict.get(mapping.get("name", "Name"), "")).strip()
            org = str(row_dict.get(mapping.get("org", "Organization"), "")).strip()

            if not email or "@" not in email:
                skipped += 1
                continue
            if self.db.blacklist_has(email):
                skipped += 1
                continue

            extra = {k: v for k, v in row_dict.items() if k not in mapping.values()}
            try:
                self.db.execute("INSERT INTO recipients (sequence_id, email, name, org, extra_json, sub_pool) VALUES (?, ?, ?, ?, ?, ?)",
                    (sequence_id or "leads", email, name, org, json.dumps(extra), sub_pool or ''))
                imported += 1
            except:
                skipped += 1

        self.db.commit()
        self._log(f"Imported {imported} leads, skipped {skipped}")
        return imported, skipped

    def import_blacklist(self, emails: List[str], reason: str = "imported"):
        count = 0
        for email in emails:
            email = email.strip().lower()
            if email and "@" in email:
                self.db.blacklist_add(email, reason)
                count += 1
        self._log(f"Imported {count} blacklisted emails")
        return count

    # -- Templates --
    def _parse_draft_subject(self, subject: str):
        """Map a Gmail draft subject to (sequence_id, day). Returns None if no match."""
        s = subject
        # Specific csr-wsl-5 variants first so they don't get swallowed by generic CSR
        wsl5_patterns = [
            r"(CSR[\s\-]*WSL[\s\-]*5|CSR[\s\-]*5YEAR|CSR[\s\-]*5[\s\-]*YEAR)[\s\-]*EMAIL[\s\-]*(\d+)",
            r"(CSR[\s\-]*WSL[\s\-]*5|CSR[\s\-]*5YEAR|CSR[\s\-]*5[\s\-]*YEAR).*?DAY\s*(\d+)",
            r"(CSR[\s\-]*WSL[\s\-]*5|CSR[\s\-]*5YEAR|CSR[\s\-]*5[\s\-]*YEAR)[\s\-]*(\d+)",
        ]
        for pat in wsl5_patterns:
            m = re.search(pat, s, re.IGNORECASE)
            if m:
                num = int(m.group(2))
                return "csr-wsl-5", EMAIL_NUM_TO_DAY.get(num, num)

        # Generic CSR / SCHOOL
        m = re.search(r"(SCHOOL|CSR)[\s\-]*EMAIL[\s\-]*(\d+)", s, re.IGNORECASE)
        if not m:
            m = re.search(r"(SCHOOL|CSR).*?DAY\s*(\d+)", s, re.IGNORECASE)
        if not m:
            m = re.search(r"(SCHOOL|CSR)[\s\-]*(\d+)", s, re.IGNORECASE)
        if m:
            seq = m.group(1).lower()
            num = int(m.group(2))
            return seq, EMAIL_NUM_TO_DAY.get(num, num)
        return None

    def sync_templates(self) -> dict:
        self._log("Syncing templates from Gmail...")
        drafts = self.gmail.list_drafts(100)
        loaded = 0
        found_names = []
        skipped = []

        for d in drafts:
            subject = d.get("subject", "")
            found_names.append(subject)
            draft_id = d.get("id", "")

            parsed = self._parse_draft_subject(subject)
            if not parsed:
                skipped.append(f"No match: {subject}")
                continue
            seq, day = parsed

            if seq not in SEQUENCES:
                skipped.append(f"Unknown seq {seq}: {subject}")
                continue

            if day not in SEQUENCES[seq]["days"]:
                skipped.append(f"Invalid day {day} for {seq}: {subject}")
                continue

            # RESPECT LOCK STATUS
            if self.is_template_locked(seq, day):
                skipped.append(f"Locked: {seq.upper()} Day {day} - skipping")
                self._log(f"Skipping locked template: {seq.upper()} Day {day}")
                continue

            full = self.gmail.get_draft_full(draft_id)
            if not full:
                skipped.append(f"Failed to fetch body: {subject} (draft_id={draft_id})")
                self._log(f"WARNING: Found matching draft but could not fetch body: {subject}")
                continue

            draft_subject = (full.get("subject") or "").strip()
            draft_html = (full.get("html_body") or "").strip()
            draft_text = (full.get("text_body") or "").strip()
            if not draft_subject or not draft_html:
                skipped.append(f"Empty draft: {subject}")
                self._log(f"WARNING: Skipping empty Gmail draft: {subject}")
                continue

            # Preserve existing A/B test settings; regenerate text_body from HTML if Gmail did not provide plain text
            existing = self.db.template_get(seq, day)
            synced_text = draft_text or self.html_to_text(draft_html)
            self.db.template_put(
                seq, day, draft_subject, draft_html,
                subject_b=existing.get("subject_b") if existing else None,
                ab_test=existing.get("ab_test", 0) if existing else 0,
                ab_split=existing.get("ab_split", 0.5) if existing else 0.5,
                text_body=synced_text,
                format=existing.get("format") if existing else "html"
            )
            loaded += 1
            self._log(f"Loaded: {subject} -> {seq.upper()} Day {day}")

        missing = []
        for seq_id, cfg in SEQUENCES.items():
            for day in cfg["days"]:
                if self.db.template_get(seq_id, day) is None:
                    missing.append(f"{seq_id.upper()} Day {day}")

        self._log(f"Sync complete: {loaded} loaded, {len(skipped)} skipped, {len(missing)} missing")
        if skipped:
            for s in skipped[:10]:
                self._log(f"  Skip reason: {s}")

        return {"loaded": loaded, "missing": missing, "found_names": found_names, "skipped": skipped}

    # -- Template Locking System --
    def lock_templates(self) -> dict:
        """Lock all existing templates on the DB column. Any legacy meta locks are migrated."""
        locked = 0
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                tmpl = self.db.template_get(seq_id, day)
                if tmpl:
                    # Migrate legacy meta lock if present
                    if self.db.get_meta(f"locked_{seq_id}_{day}") == "true":
                        self.db.template_lock(seq_id, day)
                        self.db.execute("DELETE FROM meta WHERE key=?", (f"locked_{seq_id}_{day}",))
                        self.db.commit()
                    elif not self.db.template_is_locked(seq_id, day):
                        self.db.template_lock(seq_id, day)
                    locked += 1
        self._log(f"Locked {locked} templates. Sync will not overwrite locked templates.")
        return {"locked": locked}

    def unlock_template(self, seq_id: str, day: int):
        self.db.template_unlock(seq_id, day)
        self.db.execute("DELETE FROM meta WHERE key=?", (f"locked_{seq_id}_{day}",))
        self.db.commit()
        self._log(f"Unlocked {seq_id.upper()} Day {day} for updates")

    def lock_template(self, seq_id: str, day: int):
        self.db.template_lock(seq_id, day)
        self.db.execute("DELETE FROM meta WHERE key=?", (f"locked_{seq_id}_{day}",))
        self.db.commit()
        self._log(f"Locked {seq_id.upper()} Day {day}")

    def is_template_locked(self, seq_id: str, day: int) -> bool:
        return self.db.template_is_locked(seq_id, day)

    def create_missing_drafts(self) -> dict:
        created = []
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                if self.db.template_get(seq_id, day) is None:
                    tmpl = self.generate_template(seq_id, day)
                    if "error" not in tmpl:
                        self.db.template_put(seq_id, day, tmpl["subject"], tmpl["html_body"], "generated",
                                              text_body=tmpl.get("text_body"), format=tmpl.get("format", "html"))
                        try:
                            to_addr = self.brief_email or self.default_sender or "om@robopirate.in"
                            draft = self.gmail.draft_email(
                                to_addr,
                                f"[TEMPLATE] {tmpl['subject']}",
                                tmpl["html_body"],
                                tmpl.get("text_body"),
                                sender=self.default_sender,
                                format=tmpl.get("format", "html")
                            )
                            created.append(f"{seq_id.upper()} Day {day}")
                            self._log(f"Created draft for {seq_id.upper()} Day {day} — review in Gmail")
                        except Exception as e:
                            self._log(f"DB saved but Gmail draft failed for {seq_id.upper()} Day {day}: {e}")
        return {"created": created, "count": len(created)}

    def get_template_status(self) -> dict:
        status = {}
        for seq_id in SEQUENCES:
            status[seq_id] = {}
            for day in SEQUENCES[seq_id]["days"]:
                tmpl = self.db.template_get(seq_id, day)
                locked = self.is_template_locked(seq_id, day)
                if tmpl:
                    source = tmpl.get("source") or "unknown"
                    subject = (tmpl.get("subject") or "").strip()
                    body = (tmpl.get("html_body") or "").strip()
                    status[seq_id][day] = {
                        "exists": True,
                        "empty": not subject or not body,
                        "locked": locked,
                        "source": source,
                        "subject": tmpl["subject"][:60] if subject else "(empty subject)",
                        "subject_b": tmpl.get("subject_b", ""),
                        "ab_test": bool(tmpl.get("ab_test", 0)),
                        "ab_split": tmpl.get("ab_split", 0.5)
                    }
                else:
                    status[seq_id][day] = {
                        "exists": False,
                        "empty": True,
                        "locked": False,
                        "source": None,
                        "subject": None,
                        "subject_b": None,
                        "ab_test": False,
                        "ab_split": 0.5
                    }
        return status

    def get_templates(self) -> dict:
        out = {}
        for seq_id in SEQUENCES:
            out[seq_id] = {}
            for day in SEQUENCES[seq_id]["days"]:
                t = self.db.template_get(seq_id, day)
                out[seq_id][day] = t
        return out

    # -- Generate Missing Template --
    def generate_template(self, seq_id: str, day: int) -> dict:
        cfg = SEQUENCES.get(seq_id)
        if not cfg:
            return {"error": "Invalid sequence"}

        assets = cfg.get("assets", {}).get(day, {})
        persona = cfg.get("persona", "school")

        content_html = self._generate_content(seq_id, day, assets)
        content_text = self._generate_text_content(seq_id, day, assets)
        subject = self._generate_subject(seq_id, day)
        preheader = PREHEADERS.get(seq_id, {}).get(day, "") if REWRITTEN_TEMPLATES_AVAILABLE else ""

        html = HTML_TEMPLATE.replace("{body}", content_html).replace("{preheader}", preheader)

        return {
            "subject": subject,
            "html_body": html,
            "text_body": content_text,
            "format": "html",
            "seq_id": seq_id,
            "day": day,
            "assets_used": list(assets.keys())
        }

    def _generate_subject(self, seq_id: str, day: int) -> str:
        if REWRITTEN_TEMPLATES_AVAILABLE and seq_id in REWRITTEN_SUBJECTS:
            return REWRITTEN_SUBJECTS[seq_id].get(day, f"RoboPirate {seq_id.upper()} - Day {day}")
        subjects = {
            "school": {
                1: "{{SCHOOL_NAME}} — Transform Your School with Hands-On STEM Labs",
                3: "{{SCHOOL_NAME}} — NEP 2020 Compliance: Is Your School Ready?",
                5: "{{PRINCIPAL_NAME}}, See How {{SCHOOL_NAME}} Can Lead STEM Education",
                7: "{{SCHOOL_NAME}} — Join 85+ Schools Already Using WSL",
                10: "{{PRINCIPAL_NAME}}, Final Call: WSL Subscription Plans for {{SCHOOL_NAME}}"
            },
            "csr": {
                1: "{{COMPANY_NAME}} — CSR Impact: 65,000+ Students Reached",
                3: "{{COMPANY_NAME}} — Schedule VII Alignment + STEM Education",
                5: "{{CSR_HEAD_NAME}}, Sangli Success Story for {{COMPANY_NAME}}",
                7: "{{COMPANY_NAME}} — FY Budget Planning: STEM Investment ROI",
                10: "{{CSR_HEAD_NAME}}, Partner with RoboPirate: Company Profile for {{COMPANY_NAME}}"
            },
            "csr-wsl-5": {
                1: "{{COMPANY_NAME}} — A 5-Year STEM Lab Where You Fund Only Year 1",
                3: "{{CSR_HEAD_NAME}}, We Already Did This — First WE Smart Lab, Full Academic Year, Government School",
                5: "{{CSR_HEAD_NAME}}, The Job Your CSR Creates — 1 Trainer, 5 Years, Trained from Underprivileged Background",
                7: "{{COMPANY_NAME}} — The Math: Rs.12L CSR + Rs.28L Government = 400 Students x 5 Years",
                10: "{{CSR_HEAD_NAME}}, Final Call — FY 2026-27 Budget Window + 90-Day Launch Plan"
            }
        }
        return subjects.get(seq_id, {}).get(day, f"RoboPirate {seq_id.upper()} - Day {day}")

    def _generate_content(self, seq_id: str, day: int, assets: dict) -> str:
        if seq_id == "school":
            return _new_school_content(day, assets)
        elif seq_id == "csr-wsl-5":
            return _new_csr_wsl5_content(day, assets)
        elif seq_id == "csr":
            return self._generate_csr_content(day, assets)
        else:
            return self._generate_csr_content(day, assets)

    def _generate_school_content(self, day: int, assets: dict) -> str:
        """Legacy school HTML generator; active routing now uses rewritten_email_templates."""
        a = assets
        contents = {
            1: f"""<p>Dear Principal,</p>
<p>Quick question: Are your students getting hands-on time with robotics, drones, and AI this year?</p>
<p>We've set up <strong>85+ WE Smart Labs</strong> across <strong>6 states</strong>. Schools like <strong>Veer Baji Prabhu Vidyalay</strong> (Sangli) started with a single room and now have students winning state-level competitions.</p>
<p>Everything's included — lab setup, 120+ kits, trained teacher, NEP curriculum, LMS. You just open the door.</p>
<p>Worth a 15-minute call to explore?</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<p style="font-size:12px;color:#8B949E;margin-bottom:8px;">Resources:</p>
<a href="{a.get('brochure','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 WSL Program PDF</a>
<a href="{a.get('video_wsl','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 See a Lab in Action</a>
<a href="{a.get('video_abp','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📺 ABP News Coverage</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Latest on Instagram</a>
</div>""",

            3: f"""<p>Dear Principal,</p>
<p>With NEP 2020 now in full implementation, schools across India are racing to comply with experiential learning and coding mandates from Class 6.</p>
<p><strong>The question is:</strong> Will your school lead this change or play catch-up?</p>
<p>WE Smart Lab provides:</p>
<ul>
<li>Ready-to-deploy STEM labs</li>
<li>NEP-aligned curriculum (grades 1-10)</li>
<li>Full-time trained teacher</li>
<li>Progress tracking dashboards</li>
<li>Quarterly reports + annual exhibition</li>
</ul>
<p>Let's discuss how {{SCHOOL_NAME}} can be NEP-ready this academic year.</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('video_abp','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;">📺 ABP News Coverage</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;">📱 Instagram</a>
</div>""",

            5: f"""<p>Dear Principal,</p>
<p>Let me share a story that might resonate with you.</p>
<p><strong>Veer Baji Prabhu Vidyalay</strong> — a school much like yours — partnered with us in 2024-25. Today, their students have built 12+ working robots, participated in state-level competitions, and seen measurable improvement in science engagement.</p>
<p>We develop detailed reports over every child — tracking attendance, project completion, competition results, and confidence growth. <strong>Prajwal</strong> (a specimen student from our program) went from back-row silence to building an obstacle-avoidance robot in six months. That's the kind of transformation we document.</p>
<p>Your school could be our next success story.</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<p style="font-size:12px;color:#8B949E;margin-bottom:8px;">See the impact:</p>
<a href="{a.get('report_vbv','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📊 Veer Baji Report</a>
<a href="{a.get('video_star','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 Student Star Video</a>
<a href="{a.get('folder_vbv','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📁 Full Folder</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            7: f"""<p>Dear Principal,</p>
<p>You're not alone in this journey. <strong>85+ schools</strong> across Maharashtra, Karnataka, Gujarat, and more have already chosen WE Smart Lab.</p>
<p>We deliver, and we always deliver. We don't let people down. Every single lab we've committed to is running, every single trainer is certified, every single school is seeing results.</p>
<p>Ready to join them?</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('profile','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 Company Profile</a>
<a href="{a.get('video_abp','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📺 ABP News</a>
<a href="{a.get('video_star','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 Student Star</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            10: f"""<p>Dear Principal,</p>
<p>I won't keep emailing you about this. You've got a school to run and I respect that.</p>
<p>But if you're even a little curious about what a WE Smart Lab could do for {{SCHOOL_NAME}}, I'll make time for a 10-minute call. No pitch, just show-and-tell.</p>
<p>We've prepared flexible subscription plans for schools of all sizes. Every plan includes: complete lab setup, 120+ DIY kits, full-time trained teacher, NEP 2020 + NCF aligned curriculum, LMS portal, assessments, and ongoing support.</p>
<p>If not, I genuinely wish you a great academic year.</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('plans','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 Plans & Pricing</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>"""
        }
        return contents.get(day, f"<p>Template content for Day {day}</p>")

    def _generate_csr_content(self, day: int, assets: dict) -> str:
        a = assets
        contents = {
            1: f"""<p>Dear CSR Head,</p>
<p>Your CSR budget has the power to change <strong>thousands</strong> of young lives. The question is: where will it create the most lasting impact?</p>
<p>RoboPirate's <strong>WE Smart Lab</strong> sets up fully managed STEAM/AI Smart Labs inside schools across India. As of now, we've reached <strong>65,000+ students</strong> across <strong>6 states</strong> with <strong>85+ labs</strong> delivered through strategic CSR partnerships.</p>
<p>We deliver, and we always deliver. We don't let people down.</p>
<p>Would you be open to exploring how your CSR mandate can create measurable STEM impact?</p>
<p>Best regards,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<p style="font-size:12px;color:#8B949E;margin-bottom:8px;">See the evidence:</p>
<a href="{a.get('report_sangli1','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📊 Sangli Impact Report</a>
<a href="{a.get('video_abp','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📺 ABP News Coverage</a>
<a href="{a.get('video_sangli','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 Sangli Program Video</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            3: f"""<p>Dear CSR Head,</p>
<p>Schedule VII of the Companies Act explicitly supports:</p>
<ul>
<li>Education (item ii)</li>
<li>Skill development (item x)</li>
<li>Rural development (item xii)</li>
</ul>
<p>WE Smart Lab aligns perfectly with all three. We don't just set up labs — we create sustainable STEM ecosystems that run for years.</p>
<p>We deliver and we always deliver. We don't let people down.</p>
<p>Best regards,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('report_sangli1','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📊 Sangli Impact Report</a>
<a href="{a.get('brochure','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 Brochure</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            5: f"""<p>Dear CSR Head,</p>
<p>Numbers tell stories, but faces tell them better.</p>
<p><strong>Sangli District — WE Smart Lab Impact (2024-26):</strong></p>
<ul>
<li>15 schools equipped with fully managed STEAM/AI labs</li>
<li>4,500+ students trained in robotics, coding, AI & IoT</li>
<li>87% teacher satisfaction rate</li>
<li>3 students won state-level competitions</li>
<li>1.5L+ student projects completed</li>
</ul>
<p>We also teach in <strong>Baalgruh</strong> (children's homes) and run workshops for <strong>divyang students</strong>. The 2nd stage of our Sangli expansion is now live — training trainers from underprivileged backgrounds who go on to teach 200+ students each.</p>
<p>See our work on Instagram: <a href="https://www.instagram.com/p/DSSIy7nglXc/" style="color:#59ced9;">Baalgruh</a> | <a href="https://www.instagram.com/p/DTDBcsdk9FI/" style="color:#59ced9;">Veer Baji Workshop</a> | <a href="https://www.instagram.com/p/DMhEDutOrl-/" style="color:#59ced9;">Sangli Divyang 1st Workshop</a></p>
<p>This could be your company's legacy.</p>
<p>Best regards,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('report_sangli2','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📊 Sangli Report 2</a>
<a href="{a.get('report_vbv','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📊 Veer Baji Report</a>
<a href="{a.get('video_star','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 Student Star Video</a>
<a href="{a.get('folder_sangli','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📁 Sangli Folder</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            7: f"""<p>Dear CSR Head,</p>
<p>FY 2026-27 budget season is here — this is when CSR allocations are locked. Where will your CSR rupees create the most impact?</p>
<p>Consider the WE Smart Lab model:</p>
<ul>
<li>Setup cost: Rs.2.5L – 8L per school (one-time, based on tier)</li>
<li>Annual program cost: Rs.7L per school (CSR School Model)</li>
<li>Cost per student impacted: Under Rs.500/year</li>
<li>Tax benefits: 100% deductible under Companies Act 2013 Schedule VII</li>
<li>Full compliance documentation + quarterly impact reports included</li>
</ul>
<p>Partial adoption works too. They can take 3 schools, or 4, or 10. Every child reached is a life changed.</p>
<p>Let's discuss a pilot program for Q1.</p>
<p>Best regards,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('plans','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 Plans & Pricing</a>
<a href="{a.get('video_wsl','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 WSL Video</a>
<a href="{a.get('video_abp','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📺 ABP News</a>
<a href="{a.get('video_sangli','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">🎥 Sangli Video</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>""",

            10: f"""<p>Dear CSR Head,</p>
<p>This is my last email for FY 2026-27 planning. I respect your time and your decision.</p>
<p>Let me tell you a story.</p>
<p>There was a boy named <strong>Prajwal</strong> in one of our government schools. Quiet, always sitting in the back row. The kind of child teachers forget to call on. We set up a WE Smart Lab in his school — not a big one, just the basics. A few kits, a trainer who cared, and drone access.</p>
<p>Six months later, Prajwal had built a working obstacle-avoidance robot. Not from a kit manual — from his own design. His teachers showed us the report we develop over children like him. The data was clear: attendance up, science scores up, but more than that — he asked questions now. He stood in the front row.</p>
<p>That's the imprint I want to leave. Not a sales pitch. Not a begging letter. Just this: your CSR budget can create Prajwals. One at a time, or a hundred at a time. The math works either way.</p>
<p>If this resonates, you know where to find me. If not, I genuinely wish you and your team the very best this fiscal year.</p>
<p>Warmly,<br>Robo Pirate team<br>RoboPirate</p>
<div style="margin-top:20px;padding-top:15px;border-top:1px solid #2a2a4e;">
<a href="{a.get('profile','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📄 Company Profile</a>
<a href="{a.get('kits','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-right:8px;margin-bottom:8px;">📦 Sample Kits</a>
<a href="{a.get('video_ig','#')}" style="display:inline-block;background:#59ced9;color:#0A1628;padding:8px 16px;border-radius:6px;text-decoration:none;font-weight:bold;font-size:12px;margin-bottom:8px;">📱 Instagram</a>
</div>"""
        }
        return contents.get(day, f"<p>Template content for Day {day}</p>")

    def _generate_text_content(self, seq_id: str, day: int, assets: dict) -> str:
        """Generate plain text version of email content for multipart emails."""
        if seq_id == "school":
            return _new_school_text_content(day, assets)
        elif seq_id == "csr-wsl-5":
            return _new_csr_wsl5_text_content(day, assets)
        elif seq_id == "csr":
            return self._generate_csr_text_content(day, assets)
        else:
            return self._generate_csr_text_content(day, assets)

    def _generate_school_text_content(self, day: int, assets: dict) -> str:
        """Legacy school plain-text generator; active routing now uses rewritten_email_templates."""
        a = assets
        contents = {
            1: f"""Dear Principal,

Imagine your students building robots, coding drones, and exploring AI — all within your school walls. For the 2026-27 academic year, this is no longer optional.

WE Smart Lab by RoboPirate brings cutting-edge STEAM/AI education to Indian schools. We're already in 85+ labs across 6 states, impacting 65,000+ students.

Everything is included — lab setup, 120+ DIY kits, full-time trained teacher, NEP 2020 aligned curriculum, LMS portal, and ongoing support. Schools simply open the door; we handle the rest.

Would you be open to a 15-minute call to discuss how WSL can transform your school?

Best regards,
Robo Pirate team
RoboPirate · WSL Initiative
robopirate.in

---
Resources:
📄 Brochure: {a.get('brochure', 'Available on request')}
🎥 WSL Video: {a.get('video_wsl', 'Available on request')}
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            3: f"""Dear Principal,

With NEP 2020 now in full implementation and the 2026-27 academic year approaching, schools across India are racing to comply with experiential learning and coding mandates from Class 6.

The question is: Will your school lead this change or play catch-up?

WSL provides:
• Ready-to-deploy STEM labs
• NEP-aligned curriculum
• Teacher training programs
• Progress tracking dashboards

Let's discuss how your school can be NEP-ready this academic year.

Best regards,
Robo Pirate team
RoboPirate · WSL Initiative

---
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            5: f"""Dear Principal,

Let me share a story that might resonate with you.

Veer Baji Prabhu Vidyalay — a school much like yours — partnered with us in 2024-25 through our WE Smart Lab program. Today, their students have built 12+ working robots, participated in state-level competitions, and seen measurable improvement in science engagement.

Your school could be our next success story.

Best regards,
Robo Pirate team
RoboPirate · WSL Initiative

---
📊 Impact Report (Veer Baji): {a.get('report_vbv', 'Available on request')}
🎥 Student Star Video: {a.get('video_star', 'Available on request')}
📁 Full Folder: {a.get('folder_vbv', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            7: f"""Dear Principal,

You're not alone in this journey. 85+ schools across Maharashtra, Karnataka, Gujarat, and more have already chosen WSL.

Ready to join them?

Best regards,
Robo Pirate team
RoboPirate · WSL Initiative

---
📄 Company Profile: {a.get('profile', 'Available on request')}
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
🎥 Student Star Video: {a.get('video_star', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            10: f"""Dear Principal,

This is my final email for the 2026-27 academic year planning. With admissions season approaching, I don't want your students to miss this opportunity.

We've prepared flexible WE Smart Lab subscription plans for schools of all sizes. Every plan includes: complete lab setup, 120+ DIY kits, full-time trained teacher, NEP 2020 + NCF aligned curriculum, LMS portal, assessments, and ongoing support.

If now isn't the right time, I understand. But if you're even slightly curious, let's have a 10-minute conversation. No obligation.

Best regards,
Robo Pirate team
RoboPirate · WSL Initiative

---
📄 Plans & Pricing: {a.get('plans', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
"""
        }
        return contents.get(day, f"Template content for Day {day}")

    def _generate_csr_text_content(self, day: int, assets: dict) -> str:
        a = assets
        contents = {
            1: f"""Dear CSR Head,

Your CSR budget has the power to change thousands of young lives.

RoboPirate's WE Smart Lab sets up fully managed STEAM/AI Smart Labs inside schools across India. As of now, we've reached 65,000+ students across 6 states with 85+ labs delivered through strategic CSR partnerships.

Would you be open to exploring how your CSR mandate can create measurable STEM impact?

Best regards,
Robo Pirate team
RoboPirate

---
📊 Sangli Impact Report: {a.get('report_sangli1', 'Available on request')}
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
🎥 Sangli Video: {a.get('video_sangli', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            3: f"""Dear CSR Head,

Schedule VII of the Companies Act explicitly supports:
• Education (item ii)
• Skill development (item x)
• Rural development (item xii)

WSL aligns perfectly with all three.

We deliver and we always deliver. We don't let people down.

Best regards,
Robo Pirate team
RoboPirate

---
📊 Sangli Impact Report: {a.get('report_sangli1', 'Available on request')}
📄 Brochure: {a.get('brochure', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            5: f"""Dear CSR Head,

Numbers tell stories better than words.

Sangli District Phase 2 Results — WE Smart Lab Impact (Delivered 2025-26):
• 15 schools equipped with fully managed STEAM/AI labs
• 4,500+ students trained in robotics, coding, AI & IoT
• 87% teacher satisfaction rate
• 3 students won state-level competitions
• 1.5L+ student projects completed across all programs

This could be your company's legacy.

Best regards,
Robo Pirate team
RoboPirate

---
📊 Sangli Report 2: {a.get('report_sangli2', 'Available on request')}
📊 Veer Baji Report: {a.get('report_vbv', 'Available on request')}
🎥 Student Star Video: {a.get('video_star', 'Available on request')}
📁 Sangli Folder: {a.get('folder_sangli', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            7: f"""Dear CSR Head,

FY 2026-27 budget season is here — this is when CSR allocations are locked. Where will your CSR rupees create the most impact?

Consider the WE Smart Lab model:
• Setup cost: Rs.2.5L – 8L per school (one-time, based on tier)
• Annual program cost: Rs.7L per school (CSR School Model)
• Cost per student impacted: Under Rs.500/year
• Tax benefits: 100% deductible under Companies Act 2013 Schedule VII
• Full compliance documentation + quarterly impact reports included

They can take 3/4 schools or something — partial adoption is absolutely possible. Every child reached is a life changed.

Let's discuss a pilot program for Q1.

Best regards,
Robo Pirate team
RoboPirate

---
📄 Plans & Pricing: {a.get('plans', 'Available on request')}
🎥 WSL Video: {a.get('video_wsl', 'Available on request')}
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
🎥 Sangli Video: {a.get('video_sangli', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            10: f"""Dear CSR Head,

This is my final outreach for FY 2026-27 planning. With budgets being locked, I respect your time and decision.

Let me tell you a story.

There was a boy named Prajwal in one of our government schools. Quiet, always sitting in the back row. The kind of child teachers forget to call on. We set up a WE Smart Lab in his school — not a big one, just the basics. A few kits, a trainer who cared, and drone access.

Six months later, Prajwal had built a working obstacle-avoidance robot. Not from a kit manual — from his own design. His teachers showed us the report we develop over children like him. The data was clear: attendance up, science scores up, but more than that — he asked questions now. He stood in the front row.

That's the imprint I want to leave. Not a sales pitch. Not a begging letter. Just this: your CSR budget can create Prajwals. One at a time, or a hundred at a time. The math works either way.

If this resonates, you know where to find me. If not, I wish you and your team the very best this fiscal year.

Warmly,
Robo Pirate team
RoboPirate

---
📄 Company Profile: {a.get('profile', 'Available on request')}
📦 Sample Kits: {a.get('kits', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
"""
        }
        return contents.get(day, f"Template content for Day {day}")

    def _generate_csr_wsl5_text_content(self, day: int, assets: dict) -> str:
        a = assets
        contents = {
            1: f"""Dear CSR Head,

What if your CSR budget could fund a 5-year STEM lab — and you only pay for Year 1?

That's the WE Smart Lab 5-Year Model:
• Year 1: CSR funds the lab setup + first year operations (Rs.12L)
• Years 2-5: Government/Municipal funds take over through our PMC proposal
• Result: 400 students × 5 years = 2,000 lives changed

We handle everything — setup, trainer, curriculum, reporting. You fund Year 1, we make it self-sustaining.

Would you be open to a 15-minute call to see how this works?

Best regards,
Robo Pirate team
RoboPirate

---
📊 Veer Baji Report: {a.get('report_vbv', 'Available on request')}
📄 Brochure: {a.get('brochure', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            3: f"""Dear CSR Head,

We already did this. First WE Smart Lab. Full academic year. Government school.

The trainer we placed — from an underprivileged background himself — is now certified and training 200+ students. The school principal called last week to ask when we can expand to their secondary wing.

This isn't theory. It's already happening.

Best regards,
Robo Pirate team
RoboPirate

---
📊 Veer Baji Report: {a.get('report_vbv', 'Available on request')}
📺 ABP News Coverage: {a.get('video_abp', 'Available on request')}
🎥 Student Star Video: {a.get('video_star', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            5: f"""Dear CSR Head,

The job your CSR creates:

1 Trainer. 5 Years. Trained from underprivileged background.

That's not just a job. That's a career ladder. That's a family lifted. That's a community seeing what's possible.

We don't just place trainers. We train them at our Baner HQ, certify them, and support them for 5 years. The report we develop over each child tracks everything — attendance, engagement, project completion, competition results.

This is the kind of CSR impact that gets talked about in annual reports.

Best regards,
Robo Pirate team
RoboPirate

---
🎥 WSL Video: {a.get('video_wsl', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            7: f"""Dear CSR Head,

The math:

Rs.12L CSR + Rs.28L Government = 400 Students × 5 Years

That's Rs.40L total investment. Rs.20 per student per year.

For context: One corporate off-site costs more than this. One conference booth costs more than this.

But this — this changes 2,000 lives over 5 years. This creates a STEM culture in a government school that lasts decades. This is the kind of ROI no spreadsheet can capture.

Partial adoption works too. They can take 3 schools, or 4, or 10. Every child reached is a life changed.

Best regards,
Robo Pirate team
RoboPirate

---
📄 Brochure: {a.get('brochure', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
""",

            10: f"""Dear CSR Head,

This is my final email for FY 2026-27 planning.

I've shared the numbers, the stories, the math. Now I just want to leave you with this:

We deliver. We always deliver. We don't let people down.

85+ labs. 65,000+ students. 6 states. Government schools, private schools, CSR-funded, self-funded — every single one delivered. Every single one running.

If you want to see what we offer in detail: {a.get('plans', 'Available on request')}

If this resonates, you know where to find me. If not, I genuinely wish you the very best this fiscal year.

Warmly,
Robo Pirate team
RoboPirate

---
📄 Company Profile: {a.get('profile', 'Available on request')}
📱 Instagram: {a.get('video_ig', 'Available on request')}
"""
        }
        return contents.get(day, f"Template content for Day {day}")

    def save_generated_template(self, seq_id: str, day: int, create_draft: bool = True) -> bool:
        if self.is_template_locked(seq_id, day):
            self._log(f"Template {seq_id.upper()} Day {day} is locked — skipping generation")
            return False
        template = self.generate_template(seq_id, day)
        if "error" in template:
            self._log(f"Failed to generate {seq_id.upper()} Day {day}: {template['error']}")
            return False

        self.db.template_put(seq_id, day, template["subject"], template["html_body"], "generated",
                              text_body=template.get("text_body"), format=template.get("format", "html"))

        if not create_draft:
            self._log(f"Generated {seq_id.upper()} Day {day} template (DB only)")
            return True

        try:
            draft = self.gmail.draft_email(
                self.default_sender,
                f"[TEMPLATE] {template['subject']}",
                template["html_body"],
                template.get("text_body"),
                sender=self.default_sender,
                format=template.get("format", "html")
            )
            self._log(f"Generated {seq_id.upper()} Day {day} template + Gmail draft created")
            return True
        except Exception as e:
            self._log(f"Saved to DB but Gmail draft failed: {e}")
            return True

    def validate_templates(self, seq_id: str = None, auto_repair: bool = True) -> dict:
        """Check that every configured sequence/day has a non-empty template.
        Optionally regenerate missing/empty templates from built-in content.
        Returns {"ok": bool, "repaired": [...], "failed": [...], "details": [...]}
        """
        targets = [(seq_id, SEQUENCES[seq_id])] if seq_id and seq_id in SEQUENCES else SEQUENCES.items()
        repaired = []
        failed = []
        details = []
        ok = True

        for sid, cfg in targets:
            for day in cfg["days"]:
                tmpl = self.db.template_get(sid, day)
                subject = (tmpl.get("subject") or "").strip() if tmpl else ""
                body = (tmpl.get("html_body") or "").strip() if tmpl else ""
                if tmpl and subject and body:
                    details.append({"seq_id": sid, "day": day, "status": "ok"})
                    continue

                status = "missing" if not tmpl else ("empty_subject" if not subject else "empty_body")
                details.append({"seq_id": sid, "day": day, "status": status})
                if auto_repair:
                    self._log(f"[TemplateHealth] Repairing {sid.upper()} Day {day} ({status})")
                    if self.save_generated_template(sid, day, create_draft=False):
                        repaired.append(f"{sid.upper()} Day {day}")
                        # Re-verify after repair
                        tmpl = self.db.template_get(sid, day)
                        subject = (tmpl.get("subject") or "").strip() if tmpl else ""
                        body = (tmpl.get("html_body") or "").strip() if tmpl else ""
                        if tmpl and subject and body:
                            details[-1]["status"] = "repaired"
                            continue
                        status = "repair_failed"
                    else:
                        failed.append(f"{sid.upper()} Day {day}")
                else:
                    failed.append(f"{sid.upper()} Day {day}")

                details[-1]["status"] = status
                ok = False

        return {"ok": ok, "repaired": repaired, "failed": failed, "details": details}

    def get_template_health(self) -> list:
        """Return a flat list of template health details for UI badges."""
        return self.validate_templates(auto_repair=False)["details"]

    # -- Due Recipients --
    def due_recipients(self, sequence_id: str, day: int, limit=None) -> List[Recipient]:
        cfg = SEQUENCES.get(sequence_id)
        if not cfg or day not in cfg["days"]: return []

        idx = cfg["days"].index(day)
        if idx == 0:
            sql = """SELECT r.* FROM recipients r WHERE r.sequence_id=?
                AND NOT EXISTS (SELECT 1 FROM sends s WHERE s.recipient_id=r.id)
                AND NOT EXISTS (SELECT 1 FROM blacklist b WHERE b.email=r.email)
                ORDER BY r.id"""
            params = (sequence_id,)
        else:
            prev = cfg["days"][idx - 1]
            gap = day - prev
            cutoff = (datetime.now() - timedelta(days=gap)).isoformat()
            sql = """SELECT DISTINCT r.* FROM recipients r
                JOIN sends s ON s.recipient_id=r.id AND s.day=? AND s.status='sent'
                WHERE r.sequence_id=? AND s.created_at<=?
                AND NOT EXISTS (SELECT 1 FROM sends s2 WHERE s2.recipient_id=r.id AND s2.day=?)
                AND NOT EXISTS (SELECT 1 FROM blacklist b WHERE b.email=r.email)
                AND NOT EXISTS (SELECT 1 FROM sends s3 WHERE s3.recipient_id=r.id AND s3.status='replied')
                ORDER BY s.created_at"""
            params = (prev, sequence_id, cutoff, day)

        rows = self.db.execute(sql, params).fetchall()
        return [Recipient(*r) for r in rows][:limit] if limit else [Recipient(*r) for r in rows]

    # -- Render --
    def _ab_variant(self, email: str, ab_split: float) -> str:
        """Deterministically assign A/B variant based on email hash."""
        import hashlib
        h = int(hashlib.md5(email.lower().strip().encode()).hexdigest(), 16)
        return "A" if (h % 10000) / 10000.0 < ab_split else "B"

    def render(self, seq_id: str, day: int, rec: Recipient) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[str]]:
        """Render email for recipient. Returns (subject, body_html, body_text, ab_variant)."""
        tmpl = self.db.template_get(seq_id, day)
        if not tmpl:
            return None, None, None, None

        subj, body_html, body_text = tmpl["subject"] or "", tmpl["html_body"] or "", tmpl.get("text_body") or ""
        fmt = tmpl.get("format") or "html"
        variant = None
        if tmpl.get("ab_test"):
            variant = self._ab_variant(rec.email, tmpl.get("ab_split", 0.5))
            subj = tmpl["subject"] if variant == "A" else (tmpl.get("subject_b") or tmpl["subject"])

        extra = json.loads(rec.extra_json or "{}")

        placeholders = {
            "{{PRINCIPAL_NAME}}": rec.name, "{{SCHOOL_NAME}}": rec.org,
            "{{CSR_HEAD_NAME}}": rec.name, "{{COMPANY_NAME}}": rec.org,
            "{{OPENING_LINE}}": extra.get("Opening Line", extra.get("opening_line", "")),
            "{{NAME}}": rec.name, "{{ORG}}": rec.org, "{{EMAIL}}": rec.email,
        }
        for ph, val in placeholders.items():
            subj = subj.replace(ph, str(val))
            body_html = body_html.replace(ph, str(val))
            body_text = body_text.replace(ph, str(val))
        return subj, body_html, body_text, variant, fmt

    def _send_with_retry(self, to: str, subject: str, body_html: str, body_text: str = None, thread_id=None, sender: str = None, format: str = 'html', max_retries: int = 3):
        """Send via Gmail with exponential backoff for transient SSL/network errors.
        Supports optional plain text body for multipart emails."""
        if not self.gmail or not self.gmail.is_connected():
            raise Exception("Gmail not connected. Go to Settings > Google Connections.")
        sender = sender or self.default_sender
        last_err = None
        for attempt in range(max_retries):
            try:
                return self.gmail.send_email(to, subject, body_html, body_text, thread_id, sender=sender, format=format)
            except Exception as e:
                last_err = e
                err_text = str(e).lower()
                # Only retry on transient network/TLS errors
                if any(k in err_text for k in ["ssl", "wrong_version", "connection", "timeout", "temporary"]):
                    wait = 2 ** attempt
                    self._log(f"[GmailRetry] Attempt {attempt + 1}/{max_retries} failed ({e}); retrying in {wait}s...")
                    time.sleep(wait)
                    continue
                raise
        raise last_err

    # -- Send Batch (AUTO-SEND for sequences) --
    def send_batch(self, seq_id: str, day: int, limit=None, dry_run=False) -> BatchResult:
        due = self.due_recipients(seq_id, day, limit)
        if not due: return BatchResult(queued=0, sent=0)
        if dry_run: return BatchResult(queued=len(due), sent=0)

        sent = 0
        for i, rec in enumerate(due):
            subj, body_html, body_text, ab_variant, fmt = self.render(seq_id, day, rec)
            if not subj:
                self._log(f"No template for {rec.email}, skipping")
                continue
            is_plain = fmt == 'plain'
            if (is_plain and not (body_text or "").strip()) or (not is_plain and not (body_html or "").strip()):
                self._log(f"Empty body for {rec.email}, skipping")
                continue
            try:
                # Inject tracking
                send_id = self.db.campaign_queue_send(rec.id, day, subj, "pending", "pending", None, ab_variant)
                if self.tracker and self.tracker.base_url and send_id and fmt != 'plain':
                    body_html = self.tracker.inject_tracking(body_html, rec.id, None, send_id)
                msg = self._send_with_retry(rec.email, subj, body_html, body_text, sender=self.default_sender, format=fmt)
                self.db.execute("UPDATE sends SET draft_id=?, status='sent', ab_variant=? WHERE id=?",
                                (msg.get("id"), ab_variant, send_id))
                self.db.commit()
                sent += 1
                self._log(f"Sent to {rec.email}")
                time.sleep(SEND_DELAY)
            except Exception as e:
                err = str(e)
                if "quota" in err.lower() or "rate" in err.lower() or "limit" in err.lower():
                    self._log("Rate limit hit. Saving remaining to pending_resumes...")
                    for r in due[i:]:
                        rs, rb, rt, _, _ = self.render(seq_id, day, r)
                        self.db.execute(
                            "INSERT INTO pending_resumes (sequence_id, day, recipient_id, subject, status, error) VALUES (?, ?, ?, ?, ?, ?)",
                            (seq_id, day, r.id, rs or subj, "pending", err[:200])
                        )
                    self.db.commit()
                    remaining = len(due) - i
                    self._log(f"Saved {remaining} emails to pending_resumes. Type 'resume batch {seq_id} day {day}' to continue.")
                    return BatchResult(queued=len(due), sent=sent, error="quota_hit")
                self.db.execute("UPDATE sends SET status='failed' WHERE id=?", (send_id,))
                self.db.commit()
                self._log(f"Failed: {rec.email} -- {e}")
        return BatchResult(queued=len(due), sent=sent)

    # -- Trial Send --
    def trial_send(self, email: str, seq_id: str, name="", org="") -> dict:
        """Send all 5 days of a sequence to a single email with 2-minute gaps."""
        if seq_id not in SEQUENCES:
            return {"success": False, "error": f"Unknown sequence {seq_id}"}

        # Ensure templates are valid before starting
        health = self.validate_templates(seq_id, auto_repair=True)
        if not health["ok"]:
            return {
                "success": False,
                "error": f"Templates missing/empty for {seq_id.upper()} and could not be auto-repaired: {', '.join(health['failed'])}"
            }

        days = SEQUENCES[seq_id]["days"]
        # Create a temporary recipient for rendering
        rec = Recipient(id=0, sequence_id=seq_id, email=email, name=name or "CSR Head", org=org or "Company", extra_json="{}", sub_pool="")

        results = []
        for i, day in enumerate(days):
            subj, body_html, body_text, _, fmt = self.render(seq_id, day, rec)
            if not subj:
                results.append({"day": day, "status": "skipped", "error": "missing_template"})
                self._log(f"Trial Day {day}: No template, skipped")
                continue
            if (fmt == 'plain' and not body_text.strip()) or (fmt != 'plain' and not body_html.strip()):
                results.append({"day": day, "status": "skipped", "error": "empty_body"})
                self._log(f"Trial Day {day}: Template body is empty, skipped")
                continue
            try:
                # Add a small trial marker; avoid heavy HTML banners that trigger spam filters
                trial_note_html = f"<p style='font-family:Arial,sans-serif;font-size:13px;color:#555;margin-bottom:12px'>[Raj Trial] Day {day} of {len(days)} — Sequence: {seq_id.upper()}</p>"
                trial_note_text = f"[Raj Trial] Day {day} of {len(days)} — Sequence: {seq_id.upper()}\n{'-'*40}\n\n"
                if fmt == 'plain':
                    self._send_with_retry(email, f"[Raj Trial] {subj}", "", trial_note_text + (body_text or ""), format='plain')
                else:
                    self._send_with_retry(email, f"[Raj Trial] {subj}", trial_note_html + body_html, trial_note_text + (body_text or ""), format='html')
                self._log(f"Trial sent: {seq_id.upper()} Day {day} to {email}")
                results.append({"day": day, "status": "sent"})

                # Wait 2 minutes between sends (except after last one)
                if i < len(days) - 1:
                    self._log(f"Waiting 2 minutes before Day {days[i+1]}...")
                    time.sleep(120)
            except Exception as e:
                err_text = str(e)
                self._log(f"Trial failed Day {day}: {e}")
                results.append({"day": day, "status": "failed", "error": err_text})
                break

        sent_count = sum(1 for r in results if r["status"] == "sent")
        failed = [r for r in results if r["status"] != "sent"]
        return {
            "success": sent_count > 0 or not failed,
            "sent": sent_count,
            "total": len(days),
            "results": results,
            "error": failed[0]["error"] if failed and sent_count == 0 else None
        }

    # -- Test Send --
    def test_send(self, email: str, seq_id: str, day: int, format: str = None, subject: str = None, body: str = None) -> bool:
        if seq_id not in SEQUENCES or day not in SEQUENCES[seq_id]["days"]:
            self._log("Invalid sequence or day")
            return False
        health = self.validate_templates(seq_id, auto_repair=True)
        if not health["ok"]:
            self._log(f"Template health check failed: {', '.join(health['failed'])}")
            return False
        tmpl = self.db.template_get(seq_id, day)
        if not tmpl or not (tmpl.get("subject") or "").strip():
            self._log("No valid template found")
            return False
        # The editor toggle overrides the saved format so the test matches what is on screen
        fmt = format if format in ("html", "plain") else (tmpl.get("format") or "html")
        try:
            body_text = tmpl.get("text_body") or self.html_to_text(tmpl.get("html_body", ""))
            body_html = tmpl.get("html_body", "")
            if body and body.strip():
                # Send exactly what the editor shows for the selected format
                if fmt == 'plain':
                    body_text = body
                else:
                    body_html = body
                    body_text = self.html_to_text(body)
            subj = (subject or "").strip() or tmpl["subject"]
            if fmt == 'plain':
                if not body_text.strip():
                    self._log("No plain text body found")
                    return False
                self._send_with_retry(email, f"[Raj Test] {subj}", "", body_text, format='plain')
            else:
                if not body_html.strip():
                    self._log("No HTML body found")
                    return False
                self._send_with_retry(email, f"[Raj Test] {subj}", body_html, body_text, format='html')
            self._log(f"Test sent to {email} ({fmt})")
            return True
        except Exception as e:
            self._log(f"Test failed: {e}")
            return False

    # -- Summary --
    def get_summary(self) -> dict:
        return self.db.get_dashboard_summary()

    def get_catch_up(self) -> List[dict]:
        catch = []
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                due = self.due_recipients(seq_id, day)
                if due:
                    overdue = 0
                    if day != 1:
                        prev = SEQUENCES[seq_id]["days"][SEQUENCES[seq_id]["days"].index(day) - 1]
                        oldest = self.db.execute("SELECT MIN(created_at) FROM sends s JOIN recipients r ON r.id=s.recipient_id WHERE r.sequence_id=? AND s.day=?", (seq_id, prev)).fetchone()[0]
                        if oldest:
                            expected = datetime.fromisoformat(oldest) + timedelta(days=(day - prev))
                            overdue = max(0, (datetime.now() - expected).days)
                    catch.append({"sequence": seq_id, "day": day, "count": len(due), "overdue_by_days": overdue})
        return catch

    # -- Batch Pipeline --
    def get_batch_pipeline(self, batch_id: int) -> dict:
        return self.db.batch_get_pipeline(batch_id)

    def get_all_batch_pipelines(self, sequence_id: str = None) -> list:
        return self.db.batch_get_all_pipelines(sequence_id)

    # -- POOL METHODS (NEW) --
    def get_pool(self, sequence_id: str, sub_pool: str = None, limit: int = None) -> list:
        return self.db.get_pool(sequence_id, sub_pool, limit)

    def get_pool_count(self, sequence_id: str, sub_pool: str = None) -> int:
        return self.db.get_pool_count(sequence_id, sub_pool)

    def create_batch_from_pool(self, name: str, sequence_id: str = None, batch_size: int = 10,
                                sub_pool: str = None, day_offset: int = 1, scheduled_at: str = None,
                                timezone: str = 'Asia/Kolkata', send_rate: int = 0,
                                stagger_minutes: int = 2, source_sequence: str = None) -> dict:
        # sequence_id is the target campaign sequence. source_sequence is the pool to pull from.
        target_seq = sequence_id if sequence_id in SEQUENCES else None
        pool_seq = source_sequence or target_seq or sequence_id or "leads"
        pool_count = self.get_pool_count(pool_seq, sub_pool)
        if pool_count == 0:
            return {"success": False, "error": f"No unbatched leads in {'generic' if not sub_pool else sub_pool} pool"}

        batch_seq = target_seq if target_seq else "unassigned"
        batch_id, error = self.db.batch_from_pool(
            name=name,
            sequence_id=batch_seq,
            source_sequence=source_sequence,
            batch_size=batch_size,
            sub_pool=sub_pool,
            day_offset=day_offset,
            scheduled_at=scheduled_at,
            timezone=timezone,
            send_rate=send_rate,
            stagger_minutes=stagger_minutes
        )

        if error:
            return {"success": False, "error": error}

        # Make sure recipients know their real sequence so templates render correctly
        seq_result = None
        if target_seq:
            try:
                seq_result = self.db.assign_sequence_to_batch(batch_id, target_seq)
                if seq_result.get("skipped", 0):
                    self._log(f"[POOL] Batch created; {seq_result['assigned']} assigned to {target_seq}, {seq_result['skipped']} skipped (duplicate email in target sequence)")
            except Exception as e:
                self._log(f"[POOL] Batch created but sequence assignment failed: {e}")

        actual_size = self.db.batch_count_recipients(batch_id)
        self._log(f"[POOL] Created batch '{name}' ({batch_seq.upper()}) D{day_offset} with {actual_size}/{batch_size} leads ({pool_count} available)")
        return {
            "success": True,
            "batch_id": batch_id,
            "name": name,
            "sequence_id": target_seq or batch_seq,
            "size": actual_size,
            "requested_size": batch_size,
            "pool_remaining": pool_count - actual_size,
            "day_offset": day_offset,
            "scheduled_at": scheduled_at,
            "sequence_assigned": seq_result.get("assigned", 0) if target_seq else actual_size,
            "sequence_skipped": seq_result.get("skipped", 0) if target_seq else 0,
        }

    # -- Blacklist --


    def assign_sequence_to_batch(self, batch_id: int, sequence_id: str) -> dict:
        """Assign a sequence to a batch and update its recipients."""
        if sequence_id not in SEQUENCES:
            return {"success": False, "error": f"Invalid sequence {sequence_id}"}
        try:
            rows = self.db.assign_sequence_to_batch(batch_id, sequence_id)
            self._log(f"Assigned sequence {sequence_id.upper()} to batch {batch_id} ({rows} rows updated)")
            return {"success": True, "rows_updated": rows}
        except Exception as e:
            self._log(f"Failed to assign sequence: {e}")
            return {"success": False, "error": str(e)}

    def delete_batch(self, batch_id) -> dict:
        """Soft-delete a batch: hide from active view and return leads to pool."""
        batch = self.db.batch_get(batch_id)
        if not batch:
            return {"success": False, "error": "Batch not found"}
        if batch.get("status") == "running":
            return {"success": False, "error": "Cannot delete a running batch. Pause it first."}
        try:
            returned = self.db.batch_soft_delete(batch_id)
            self._log(f"Deleted batch {batch_id} — {returned} leads returned to pool")
            return {"success": True, "returned": returned, "batch_id": batch_id}
        except Exception as e:
            self._log(f"Failed to delete batch {batch_id}: {e}")
            return {"success": False, "error": str(e)}

    def clone_family(self, source_family_name: str, new_family_name: str, sub_pool: str = None) -> dict:
        """Clone a campaign family: create a new Day 1 batch with the same settings and fresh leads."""
        # Find the source Day 1 batch (handles old naming Family-B1 and new naming Family-B1-D1)
        source_rows = self.db.execute("""
            SELECT * FROM batches
            WHERE (name=? OR name LIKE ?) AND day_offset=1 AND deleted_at IS NULL
            ORDER BY id DESC LIMIT 1
        """, (source_family_name, f"{source_family_name}-%")).fetchall()
        if not source_rows:
            return {"success": False, "error": f"Source family '{source_family_name}' not found"}
        source = dict(source_rows[0])
        seq_id = source["sequence_id"]
        source_batch_id = source["id"]

        # Determine source sub-pool if none provided
        if sub_pool is None:
            sub_rows = self.db.execute("""
                SELECT DISTINCT r.sub_pool
                FROM recipients r
                JOIN batch_recipients br ON r.id = br.recipient_id
                WHERE br.batch_id = ? AND r.sub_pool != ''
            """, (source_batch_id,)).fetchall()
            sub_pool = sub_rows[0][0] if sub_rows else None

        # Match source size
        source_size = self.db.batch_count_recipients(source_batch_id)
        if source_size == 0:
            return {"success": False, "error": "Source family has no leads"}

        new_name = f"{new_family_name}-D1"
        result = self.create_batch_from_pool(
            name=new_name,
            sequence_id=seq_id,
            batch_size=source_size,
            sub_pool=sub_pool,
            day_offset=1
        )
        if not result.get("success"):
            return result

        self._log(f"[CLONE] '{source_family_name}' → '{new_family_name}' ({result['size']} leads, sub-pool: {sub_pool or 'all'})")
        return {
            "success": True,
            "batch_id": result["batch_id"],
            "name": new_name,
            "sequence_id": seq_id,
            "size": result["size"],
            "sub_pool": sub_pool
        }

    def blacklist_add(self, email: str, reason: str = "manual"):
        self.db.blacklist_add(email, reason)
        self._log(f"Blacklisted: {email}")

    def blacklist_remove(self, email: str):
        self.db.blacklist_remove(email)
        self._log(f"Removed from blacklist: {email}")

    # -- Bounce Scan --
    def _check_bounce_scan(self, now: datetime):
        last = self.db.get_meta("last_bounce_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(hours=BOUNCE_INTERVAL): return
        self.scan_bounces(days_back=15)

    def scan_bounces(self, days_back: int = 1) -> dict:
        """Scan for bounces and auto-replies. Deletes processed emails from Gmail."""
        self.db.set_meta("last_bounce_scan", datetime.now().isoformat())
        last = self.db.get_meta("last_bounce_scan")
        if last:
            last_dt = datetime.fromisoformat(last)
            scan_since = max(last_dt, datetime.now() - timedelta(days=days_back))
        else:
            scan_since = datetime.now() - timedelta(days=days_back)

        after_str = scan_since.strftime("%Y/%m/%d")

        queries = [
            f"after:{after_str} (from:mailer-daemon OR from:postmaster OR from:Mail Delivery Subsystem OR from:MAILER-DAEMON)",
            f"after:{after_str} (subject:undelivered OR subject:bounce OR subject:'delivery status' OR subject:'delivery failure' OR subject:'failed delivery' OR subject:'address not found' OR subject:'recipient not found' OR subject:'Mail delivery failed' OR subject:'Returned mail')",
            f"after:{after_str} (subject:out of office OR subject:vacation OR subject:'auto reply' OR subject:'automated response' OR subject:'automatic reply' OR subject:'away from office')",
        ]

        all_msgs = []
        seen_ids = set()

        for q in queries:
            try:
                msgs = self.gmail.search_messages(q, 100)
                for m in msgs:
                    if m['id'] not in seen_ids:
                        seen_ids.add(m['id'])
                        all_msgs.append(m)
            except Exception as e:
                self._log(f"Bounce query failed: {e}")

        self._log(f"Bounce scan: {len(all_msgs)} messages to check")

        new_blacklisted = 0
        auto_reply_count = 0
        protected_count = 0
        deleted_count = 0
        skipped = 0
        processed_this_scan = set()

        for msg in all_msgs:
            subject = msg.get("subject", "").lower()
            body = msg.get("body", "") or ""
            from_addr = msg.get("from", "").lower()
            snippet = msg.get("snippet", "") or ""
            msg_id = msg["id"]

            if "robopirate" in from_addr and "mailer-daemon" not in from_addr:
                continue

            is_bounce = self._looks_like_bounce(from_addr, subject, body)
            is_auto_reply = self._is_auto_reply(subject, body)

            if is_auto_reply and not is_bounce:
                auto_reply_count += 1
                self._log(f"[AUTO-REPLY] {from_addr[:40]}: {subject[:50]}")
                self._delete_bounce_email(msg_id)
                continue

            if not is_bounce:
                self._delete_bounce_email(msg_id)
                continue

            addrs = self._extract_bounced(body) or []

            try:
                full = self.gmail.get_message_full(msg_id)
                if full:
                    full_addrs = self._extract_bounced(full.get("body", "") or "")
                    for a in full_addrs:
                        if a not in addrs:
                            addrs.append(a)
            except:
                pass

            snippet_addrs = self._extract_bounced(snippet)
            for a in snippet_addrs:
                if a not in addrs:
                    addrs.append(a)

            if not addrs:
                self._log(f"[BOUNCE] No address extracted: {subject[:60]}")
                self._delete_bounce_email(msg_id)
                continue

            for addr in addrs:
                addr = addr.lower().strip()

                if not addr or "@" not in addr:
                    continue
                if addr.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".ico", ".css", ".js")):
                    continue
                if "/" in addr or "?" in addr or "&" in addr:
                    continue
                if addr.startswith(("wght@", "size@", "color@", "font@")):
                    continue
                if self.is_protected_email(addr):
                    protected_count += 1
                    continue
                if self.db.blacklist_has(addr):
                    if addr not in processed_this_scan:
                        skipped += 1
                        processed_this_scan.add(addr)
                    continue
                if addr in processed_this_scan:
                    continue

                processed_this_scan.add(addr)

                # ── Layer 1: Verify email is in our recipient pool ──
                if not self.db.recipient_exists(addr):
                    self._log(f"[BOUNCE-SKIP] {addr} — not in recipient pool")
                    continue

                # ── Layer 2: Verify we actually sent to them ──
                if not self.db.was_sent_to(addr):
                    self._log(f"[BOUNCE-SKIP] {addr} — no send record")
                    continue

                # ── Layer 3: Classify hard vs soft bounce ──
                bounce_type, reason = self._classify_bounce(body)
                if bounce_type == "soft":
                    self._log(f"[BOUNCE-SOFT] {addr} — {reason}")
                    continue

                # ── All checks passed — blacklist with reason ──
                self.db.blacklist_add(addr, f"bounce: {reason}")
                self.db.mark_email_bounced(addr, reason)
                new_blacklisted += 1
                self._log(f"[BLACKLIST] {addr} — {reason}")
                self._notify("Bounced", f"{addr}. Blacklisted.")

            self._delete_bounce_email(msg_id)
            deleted_count += 1

        self._log(f"Bounce scan: {new_blacklisted} new blacklisted, {auto_reply_count} auto-replies, {protected_count} protected, {deleted_count} deleted, {skipped} already blacklisted")
        return {
            "new_blacklisted": new_blacklisted,
            "auto_replies": auto_reply_count,
            "protected": protected_count,
            "deleted": deleted_count,
            "skipped": skipped
        }

    def _delete_bounce_email(self, msg_id: str):
        """Delete a bounce email from Gmail."""
        try:
            self.gmail.trash_message(msg_id)
        except Exception as e:
            try:
                self.gmail.delete_message(msg_id)
            except:
                self._log(f"Could not delete bounce email {msg_id}: {e}")

    def deep_bounce_scan(self, days: int = 30) -> dict:
        """Deep scan inbox for bounce emails over last N days."""
        results = {'found': 0, 'blacklisted': 0, 'protected': 0, 'details': []}
        try:
            after_date = (datetime.now() - timedelta(days=days)).strftime("%Y/%m/%d")
            query = f"after:{after_date} (from:mailer-daemon OR from:postmaster OR 'delivery status notification' OR 'undeliverable' OR 'message not delivered')"
            messages = self.gmail.search_messages(query, max_results=200)
            if not messages:
                self._log(f"[DEEP BOUNCE SCAN] No bounce emails found in last {days} days")
                return results

            self._log(f"[DEEP BOUNCE SCAN] Checking {len(messages)} potential bounce emails (last {days} days)...")

            sent_rows = self.db.execute("SELECT DISTINCT email FROM recipients").fetchall()
            our_emails = {r[0].lower().strip() for r in sent_rows}

            for msg in messages:
                try:
                    from_addr = msg.get('from', '').lower()
                    subject = msg.get('subject', '').lower()
                    body = msg.get('body', '').lower()

                    is_mailer = any(x in from_addr for x in ['mailer-daemon', 'postmaster', 'mail delivery subsystem'])
                    is_bounce_subject = any(x in subject for x in [
                        'delivery status notification', 'undeliverable', 'permanent failure',
                        'message not delivered', 'failure notice', 'returned mail'
                    ])
                    if not (is_mailer or is_bounce_subject):
                        continue

                    results['found'] += 1

                    bounced_email = None
                    patterns = [
                        r'final-recipient:\s*rfc822;\s*([^\s<>]+)',
                        r'original-recipient:\s*rfc822;\s*([^\s<>]+)',
                        r'to:\s*([^\s<>]+@[^\s<>]+)',
                        r'does not exist[:\s]+([^\s<>]+@[^\s<>]+)',
                    ]
                    for pat in patterns:
                        m = re.search(pat, body)
                        if m:
                            bounced_email = m.group(1).strip()
                            break

                    if not bounced_email:
                        emails_in_body = re.findall(r'[\w.-]+@[\w.-]+\.\w+', body)
                        for e in emails_in_body:
                            if e.lower() in our_emails:
                                bounced_email = e
                                break

                    if not bounced_email:
                        continue

                    bounced_email = bounced_email.lower().strip()
                    bounced_email = re.sub(r"[.,;:!?)\'\"]+$", "", bounced_email)

                    if not re.match(r'^[\w.-]+@[\w.-]+\.\w+$', bounced_email):
                        continue
                    if self.db.blacklist_has(bounced_email):
                        continue
                    if self.is_protected_email(bounced_email):
                        results['protected'] += 1
                        continue
                    if bounced_email not in our_emails:
                        continue

                    # Layer 2: Verify we actually sent to them
                    if not self.db.was_sent_to(bounced_email):
                        continue

                    # Layer 3: Classify hard vs soft
                    bounce_type, reason = self._classify_bounce(body)
                    if bounce_type == "soft":
                        results['details'].append({'email': bounced_email, 'action': 'SKIPPED (soft bounce)'})
                        continue

                    self.db.blacklist_add(bounced_email, f"bounce: {reason} (deep scan {days}d)")
                    self.db.mark_email_bounced(bounced_email, reason)
                    results['blacklisted'] += 1
                    results['details'].append({'email': bounced_email, 'action': 'BLACKLISTED'})
                    self.gmail.trash_message(msg['id'])

                except Exception as e:
                    continue

            self._log(f"[DEEP BOUNCE SCAN] Complete: {results['found']} found, {results['blacklisted']} blacklisted, {results['protected']} protected")
            return results

        except Exception as e:
            self._log(f"[Engine] Deep bounce scan error: {e}")
            return results

    @staticmethod
    def is_protected_email(email: str) -> bool:
        """Check if an email is protected from blacklisting."""
        if not email:
            return False
        email = email.lower().strip()
        return email.endswith("@robopirate.in") or email == "itsomkarsinghhh@gmail.com"

    def _classify_bounce(self, body: str) -> tuple:
        """Classify bounce as hard (permanent) or soft (temporary).
        Returns (type, reason) where type is 'hard' or 'soft'.
        Only blacklists on explicit hard 5xx/2.5.x DSN codes or clear hard keywords.
        """
        if not body:
            return "soft", "unknown"
        body_lower = body.lower()

        # Check SMTP DSN status codes first (most reliable): 5.x.y = hard, 4.x.y = soft
        for m in re.finditer(r'status:\s*(\d\.\d\.\d)', body_lower):
            code = m.group(1)
            if code.startswith('5'):
                return "hard", f"SMTP {code}"
            if code.startswith('4'):
                return "soft", f"SMTP {code}"

        # Bare 3-digit SMTP codes only when explicitly tied to a delivery status
        if re.search(r'(smtp|delivery|status|code|error)\s*[:#]?\s*5\d{2}', body_lower):
            code = re.search(r'5\d{2}', body_lower).group(0)
            return "hard", f"SMTP {code}"
        if re.search(r'(smtp|delivery|status|code|error)\s*[:#]?\s*4\d{2}', body_lower):
            code = re.search(r'4\d{2}', body_lower).group(0)
            return "soft", f"SMTP {code}"

        # Hard bounce keywords (permanent failures)
        hard_keywords = [
            "user unknown", "no such user", "address does not exist",
            "invalid address", "domain not found", "mailbox unavailable",
            "recipient address rejected", "permanent failure", "does not exist",
            "unable to deliver", "delivery permanently", "not a valid",
            "host unknown", "unrouteable address", "relay access denied"
        ]
        for kw in hard_keywords:
            if kw in body_lower:
                return "hard", kw

        # Soft bounce keywords (temporary failures)
        soft_keywords = [
            "mailbox full", "quota exceeded", "temporary failure",
            "try again later", "server busy", "defer", "delayed",
            "greylisted", "temporarily rejected", "soft bounce",
            "resource temporarily unavailable"
        ]
        for kw in soft_keywords:
            if kw in body_lower:
                return "soft", kw

        # Default: unknown = soft (do not auto-blacklist ambiguous bounces)
        return "soft", "unknown"

    def _looks_like_bounce(self, from_addr: str, subject: str, body: str) -> bool:
        """Quick heuristic check if an email looks like a bounce or auto-reply."""
        from_lower = from_addr.lower()
        subj_lower = subject.lower()
        body_lower = body.lower()

        bounce_senders = [
            "mailer-daemon", "postmaster", "mail delivery subsystem",
            "daemon", "bounce", "undeliverable", "noreply"
        ]
        for sender in bounce_senders:
            if sender in from_lower:
                return True

        bounce_subjects = [
            "undelivered", "bounce", "delivery status", "delivery failure",
            "failed delivery", "address not found", "recipient not found",
            "returned mail", "mail delivery failed", "message not delivered"
        ]
        for pattern in bounce_subjects:
            if pattern in subj_lower:
                return True

        auto_subjects = [
            "out of office", "auto reply", "automated response", "automatic reply",
            "vacation", "on leave", "away from office", "abwesenheitsnotiz"
        ]
        for pattern in auto_subjects:
            if pattern in subj_lower:
                return True

        body_bounce_patterns = [
            "final-recipient", "diagnostic-code", "action: failed",
            "status:", "remote server", "smtp error", "550 ", "551 ",
            "552 ", "553 ", "554 ", "recipient address rejected",
            "user unknown", "no such user", "mailbox unavailable"
        ]
        for pattern in body_bounce_patterns:
            if pattern in body_lower:
                return True

        auto_body_patterns = [
            "auto-submitted:", "x-autoreply:", "precedence: auto_reply",
            "x-auto-response-suppress:", "i am currently out of",
            "i will be out of", "i am away", "on vacation until",
            "return on", "back on", "this is an automated"
        ]
        for pattern in auto_body_patterns:
            if pattern in body_lower:
                return True

        return False

    @staticmethod
    def _is_auto_reply(subject: str, body: str) -> bool:
        """Detect if message is an auto-reply/out-of-office/vacation response."""
        subject_lower = subject.lower()
        body_lower = body.lower()

        # Detect true auto-replies: header declaration + subject-line hints.
        # We deliberately do NOT match body phrases like "thank you for your email"
        # because real human replies often contain them.
        auto_reply_headers = ["auto-submitted", "autoreplied", "x-autoreply"]
        if any(h in body_lower or h in subject_lower for h in auto_reply_headers):
            return True

        auto_reply_subjects = [
            "out of office", "out of the office", "away from office", "on vacation",
            "on leave", "auto reply", "automated response", "automatic reply",
            "automatic response", "auto-response", "out of office reply",
            "abwesenheitsnotiz", "risposta automatica", "respuesta automatica",
            "reponse automatique", "automatikus valasz", "automatski odgovor",
            "automatisch antwoord", "automaattinen vastaus", "automatsvar",
            "i am currently out of", "i will be out of", "i am away",
            "not in office", "not available", "currently unavailable",
            "this is an automated", "this email is automatically",
            "do not reply to this", "noreply", "no reply",
            "i am on holiday", "i am on vacation", "annual leave",
            "maternity leave", "paternity leave", "sick leave",
            "traveling until", "back on", "return on", "will return",
            "limited access to email", "intermittent access",
            "email access limited", "delayed response"
        ]

        for keyword in auto_reply_subjects:
            if keyword in subject_lower:
                return True

        header_patterns = [
            "auto-submitted:", "x-autoreply:", "x-auto-response-suppress:",
            "precedence: auto_reply", "precedence: bulk",
            "x-mailer: autoreply", "x-autoresponder:",
            "vacation:", "x-vacation:", "autoreply:"
        ]
        for pattern in header_patterns:
            if pattern in body_lower:
                return True

        return False

    def _extract_original_sender(self, subject: str, body: str, msg_or_full: dict = None) -> Optional[str]:
        """Extract the original sender email from an auto-reply or bounce message."""
        patterns = [
            r"Original-From:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"From:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"Sender:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"Reply-To:\s*<?([\w.+-]+@[\w.-]+)>?",
            r"was sent by\s+([\w.+-]+@[\w.-]+)",
            r"sent by\s+([\w.+-]+@[\w.-]+)",
            r"original message was sent by\s+([\w.+-]+@[\w.-]+)",
            r"your message to\s+([\w.+-]+@[\w.-]+)",
            r"email sent to\s+([\w.+-]+@[\w.-]+)",
            r"message to\s+([\w.+-]+@[\w.-]+)\s+was",
        ]

        texts = [body]
        if msg_or_full:
            texts.append(msg_or_full.get("body", ""))
            texts.append(msg_or_full.get("snippet", ""))

        for text in texts:
            if not text:
                continue
            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    email = match.group(1).strip().strip("<>").lower()
                    if "@" in email and "mailer-daemon" not in email and "postmaster" not in email:
                        return email
            to_match = re.search(r"To:\s*<?([\w.+-]+@[\w.-]+)>?", text, re.IGNORECASE)
            if to_match:
                email = to_match.group(1).strip().strip("<>").lower()
                if "@" in email and "mailer-daemon" not in email and "postmaster" not in email:
                    return email

        return None

    @staticmethod
    def _extract_bounced(text: str) -> List[str]:
        if not text:
            return []
        addrs = []

        patterns = [
            r"Final-Recipient:\s*rfc822;\s*([\w.+-]+@[\w.-]+)",
            r"Original-Recipient:\s*rfc822;\s*([\w.+-]+@[\w.-]+)",
            r"To:\s*<([\w.+-]+@[\w.-]+)>",
            r"Your message to\s+([\w.+-]+@[\w.-]+)\s+couldn'?t be delivered",
            r"message to\s+([\w.+-]+@[\w.-]+)\s+was undeliverable",
            r"(?:was not delivered to|wasn'?t delivered to|could not be delivered to|couldn't be delivered to|failed to deliver to)\s+([\w.+-]+@[\w.-]+)",
            r"Address not found.*?(?:to|for)\s+([\w.+-]+@[\w.-]+)",
            r"^\s*<([\w.+-]+@[\w.-]+)>:?\s*$",
            r"([\w.+-]+@[\w.-]+):\s*(?:user unknown|mailbox unavailable|no such user|does not exist|mailbox full|invalid user|unknown local-part)",
            r"did not reach.*?([\w.+-]+@[\w.-]+)",
            r"address(?:es)? failed.*?([\w.+-]+@[\w.-]+)",
        ]

        for pat in patterns:
            for m in re.finditer(pat, text, re.I | re.M | re.DOTALL):
                email = m.group(1).strip().strip("<>").lower()
                if "@" in email and "mailer-daemon" not in email and "postmaster" not in email:
                    if email not in addrs:
                        addrs.append(email)

        if not addrs:
            for m in re.finditer(r"<([\w.+-]+@[\w.-]+)>", text):
                email = m.group(1).strip().lower()
                if "mailer-daemon" not in email and "postmaster" not in email:
                    if email not in addrs:
                        addrs.append(email)

        if not addrs and any(k in text.lower() for k in ["delivery", "bounce", "failed", "undelivered", "address not found", "recipient", "mailer-daemon", "postmaster"]):
            for m in re.finditer(r"[\w.+-]+@[\w.-]+", text):
                email = m.group().lower()
                if any(x in email for x in ["mailer-daemon", "postmaster", "robopirate.in", "google.com", "gmail.com", "instagram", "facebook", "twitter", "linkedin", "youtube", "2x", "3x", "1x", "wght", "size", "color", "font"]):
                    continue
                if "@" in email:
                    domain = email.split("@")[1]
                    if "." not in domain or len(domain) < 4:
                        continue
                if email not in addrs:
                    addrs.append(email)

        return addrs

    def _check_reply_scan(self, now: datetime):
        last = self.db.get_meta("last_reply_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(minutes=REPLY_INTERVAL): return
        self.scan_replies()

    def scan_replies(self, days_back: int = 3) -> int:
        """Scan inbox for replies from recipients."""
        self.db.set_meta("last_reply_scan", datetime.now().isoformat())
        after = int((datetime.now() - timedelta(days=days_back)).timestamp())

        msgs_all = self.gmail.search_messages(f"in:inbox after:{after}", 200)
        msgs_sent = self.gmail.search_messages(f"in:sent after:{after}", 100)
        msgs_re = self.gmail.search_messages(f"in:inbox subject:Re: after:{after}", 100)

        seen_ids = set()
        all_msgs = []
        for m in msgs_all + msgs_sent + msgs_re:
            if m['id'] not in seen_ids:
                seen_ids.add(m['id'])
                all_msgs.append(m)

        self._log(f"DEBUG REPLY SCAN: {len(all_msgs)} unique messages to check")

        new_count = 0
        checked_count = 0

        for msg in all_msgs:
            from_addr = email.utils.parseaddr(msg.get("from", ""))[1].lower()
            subject = msg.get("subject", "").lower()
            body = msg.get("body", "") or ""

            if "robopirate" in from_addr:
                continue

            if self._is_auto_reply(subject, body):
                continue
            if "mailer-daemon" in from_addr or "postmaster" in from_addr:
                continue

            checked_count += 1

            rows = self.db.execute("""SELECT r.id, s.id as send_id, s.draft_id, s.day
                FROM recipients r JOIN sends s ON s.recipient_id=r.id
                WHERE r.email=? AND s.status!='replied'""", (from_addr,)).fetchall()

            if not rows:
                continue

            for rec_id, send_id, draft_id, day in rows:
                if self.db.execute("SELECT 1 FROM replies WHERE message_id=?", (msg["id"],)).fetchone():
                    continue
                body = msg.get("body", "")[:2000]
                self.db.execute("""INSERT INTO replies (send_id, recipient_id, thread_id, message_id, from_addr, subject, body, received_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (send_id, rec_id, msg.get("threadId", ""), msg["id"], from_addr, msg.get("subject", ""), body, datetime.now().isoformat()))
                self.db.execute("UPDATE sends SET status='replied' WHERE id=?", (send_id,))
                new_count += 1
                self._log(f"New reply from {from_addr}: {msg.get('subject', '')[:60]}")
                break

        self.db.set_meta("last_reply_scan", datetime.now().isoformat())
        if new_count:
            self._log(f"Found {new_count} new replies (checked {checked_count} messages)")
            self._notify("New Reply", f"{new_count} replies. Check Replies tab.")
        else:
            self._log(f"No new replies found (checked {checked_count} messages)")
        return new_count

    def _check_eod(self, now: datetime):
        today_eod = now.replace(hour=EOD_HOUR, minute=0, second=0, microsecond=0)
        if now < today_eod:
            return
        last = self.db.get_meta("last_eod_run")
        if last and datetime.fromisoformat(last) >= today_eod:
            return
        self.draft_replies_eod()

    def draft_replies_eod(self) -> dict:
        import requests
        pending = self.db.execute("SELECT * FROM replies WHERE status='pending'").fetchall()
        counts = {"positive": 0, "neutral": 0, "hostile": 0, "unsubscribe": 0, "drafted": 0}

        for row in pending:
            reply = dict(row)
            reply_id = reply.get("id")
            send_id = reply.get("send_id")
            from_addr = reply.get("from_addr")
            rec = self.db.execute("""SELECT r.*, s.day, s.subject as orig_subject
                FROM recipients r JOIN sends s ON s.recipient_id=r.id WHERE s.id=?""", (send_id,)).fetchone()
            if not rec: continue
            rec = dict(rec)

            seq_id = rec.get("sequence_id", "")
            persona = SEQUENCES.get(seq_id, {}).get("persona", "school")

            context = self._build_reply_context(reply, rec)
            system = self._persona_prompt(persona)
            user = self._reply_user_prompt(context)

            try:
                r = requests.post(f"{self.ollama_url}/api/chat", json={
                    "model": "gpt-oss:20b-cloud",
                    "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
                    "stream": False
                }, timeout=120)
                content = r.json()["message"]["content"]
                m = re.search(r"\{.*\}", content, re.DOTALL)
                if not m: continue
                result = json.loads(m.group())

                sentiment = result.get("sentiment", "neutral")
                counts[sentiment] = counts.get(sentiment, 0) + 1

                if sentiment in ("hostile", "unsubscribe"):
                    self.db.blacklist_add(from_addr, f"sentiment:{sentiment}")
                    self.db.execute("UPDATE replies SET status='handled', sentiment=? WHERE id=?", (sentiment, reply_id))
                    self._log(f"Auto-blacklisted {from_addr} ({sentiment})")
                    continue

                draft_html = result.get("draft_html", "")
                sender = self.db.get_meta("default_sender") or "om@robopirate.in"
                reply_subject = reply.get("subject", "")
                reply_subject = f"Re: {reply_subject}" if reply_subject and not reply_subject.startswith("Re:") else reply_subject
                draft = self.gmail.draft_reply(reply["thread_id"], draft_html, reply_subject, to=from_addr, sender=sender)
                draft_id = draft.get("id") if draft else None
                self.db.execute("UPDATE replies SET status='drafted', sentiment=?, summary=?, draft_reply_id=?, draft_html=? WHERE id=?",
                    (sentiment, result.get("summary", ""), draft_id, draft_html, reply_id))
                counts["drafted"] += 1
                self._log(f"Drafted reply for {from_addr} ({sentiment}) -- waiting for your approval")
            except Exception as e:
                self._log(f"EOD draft failed: {e}")

        self.db.set_meta("last_eod_run", datetime.now().isoformat())
        self._log(f"EOD complete: {counts}")
        return counts

    def _persona_prompt(self, persona: str) -> str:
        return {
            "school": "You are the RoboPirate school outreach team. Warm, professional HTML emails to Indian private school principals. Never salesy.",
            "csr": "You are the RoboPirate CSR team. Formal, impact-focused emails to CSR heads. Data-driven and professional.",
            "csr-wsl-5": "You are the RoboPirate CSR team. Formal, impact-focused emails to CSR heads about the 5-year co-funded pilot model. Data-driven, employment-focused, and professional.",
        }.get(persona, "")

    def _build_reply_context(self, reply: dict, rec: dict) -> dict:
        """Gather rich context for a smart reply: recipient, original email, thread history, assets."""
        send_id = reply.get("send_id")
        seq_id = rec.get("sequence_id", "")
        day = rec.get("day", 1)
        orig_subject = rec.get("orig_subject", "")
        orig_body_text = ""
        assets = {}

        # Try to get the original sent template body and sequence assets
        if seq_id and day:
            tmpl = self.db.template_get(seq_id, day)
            if tmpl:
                orig_subject = orig_subject or tmpl.get("subject", "")
                orig_body_text = tmpl.get("text_body") or self.html_to_text(tmpl.get("html_body", ""))
            seq_cfg = SEQUENCES.get(seq_id, {})
            assets = seq_cfg.get("assets", {}).get(day, {})

        # Thread history from previous replies
        thread_id = reply.get("thread_id", "")
        history = []
        if thread_id:
            rows = self.db.execute(
                "SELECT from_addr, body, received_at, sentiment FROM replies WHERE thread_id=? ORDER BY received_at",
                (thread_id,)
            ).fetchall()
            for row in rows:
                history.append(f"{row[0]} ({row[2]}) [{row[3] or 'unknown'}]: {row[1][:400]}")

        # Recipient custom fields (designation, city, etc.)
        extra = {}
        try:
            extra = json.loads(rec.get("extra_json") or "{}")
        except Exception:
            pass

        return {
            "recipient_name": rec.get("name", ""),
            "recipient_org": rec.get("org", ""),
            "recipient_email": rec.get("email", ""),
            "recipient_extra": extra,
            "sequence": seq_id.upper(),
            "day": day,
            "original_subject": orig_subject,
            "original_body": orig_body_text[:1200],
            "reply_subject": reply.get("subject", ""),
            "reply_body": (reply.get("body") or "")[:1200],
            "thread_history": "\n---\n".join(history[-3:]),
            "assets": assets,
        }

    def _reply_user_prompt(self, context: dict) -> str:
        assets_block = "\n".join([f"- {k}: {v}" for k, v in (context.get("assets") or {}).items()]) or "No specific assets for this email."
        extra_block = "\n".join([f"- {k}: {v}" for k, v in (context.get("recipient_extra") or {}).items()]) or "No extra recipient details."
        return f"""You are drafting a reply email for RoboPirate.

Recipient: {context['recipient_name']} from {context['recipient_org']} ({context['recipient_email']})
Extra recipient details:
{extra_block}
Sequence: {context['sequence']} Day {context['day']}

Original email we sent:
Subject: {context['original_subject']}
---
{context['original_body']}
---

Recipient's reply:
Subject: {context['reply_subject']}
---
{context['reply_body']}
---

Recent thread history:
{context['thread_history'] or 'No earlier replies in thread.'}

Relevant links/assets for this sequence/day (include only if the recipient asks for them):
{assets_block}

Instructions:
- Be warm, professional, and context-aware.
- Directly address the recipient's points or questions.
- If the reply is hostile, abusive, or clearly asks to stop, set sentiment to "hostile" or "unsubscribe" and draft a brief, respectful closing.
- If the reply asks for a meeting, propose 2 short time slots and offer a calendar link.
- If the reply asks for a brochure, video, or proposal, include the most relevant asset link from the list above.
- Keep it concise (under 200 words).
- Do not be pushy or apologetic.
- Use natural Indian business English tone.
- Sign off as "Robo Pirate team, RoboPirate".
- Output valid JSON with exactly these keys: sentiment, summary, draft_html.
- sentiment must be one of: positive, neutral, hostile, unsubscribe.
- summary is a 1-sentence summary of the reply.
- draft_html is the reply body in simple HTML (just paragraphs, no full email wrapper).
"""

    def generate_reply_draft(self, reply_id: int) -> dict:
        """Generate an AI draft for a single reply. Returns sentiment, summary, draft_html."""
        import requests
        reply = self.db.execute("SELECT * FROM replies WHERE id=?", (reply_id,)).fetchone()
        if not reply:
            return {"success": False, "error": "Reply not found"}
        reply = dict(reply)

        rec = self.db.execute("""SELECT r.*, s.day, s.subject as orig_subject
            FROM recipients r JOIN sends s ON s.recipient_id=r.id WHERE s.id=?""", (reply["send_id"],)).fetchone()
        if not rec:
            return {"success": False, "error": "Recipient not found"}
        rec = dict(rec)

        seq_id = rec.get("sequence_id", "")
        persona = SEQUENCES.get(seq_id, {}).get("persona", "school")

        context = self._build_reply_context(reply, rec)
        system = self._persona_prompt(persona)
        user = self._reply_user_prompt(context)

        try:
            r = requests.post(f"{self.ollama_url}/api/chat", json={
                "model": "gpt-oss:20b-cloud",
                "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
                "stream": False
            }, timeout=120)
            content = r.json()["message"]["content"]
            m = re.search(r"\{.*\}", content, re.DOTALL)
            if not m:
                return {"success": False, "error": "AI did not return JSON"}
            result = json.loads(m.group())

            sentiment = result.get("sentiment", "neutral")
            summary = result.get("summary", "")
            draft_html = result.get("draft_html", "")

            # Optionally create Gmail draft immediately
            try:
                sender = self.db.get_meta("default_sender") or "om@robopirate.in"
                reply_subject = reply.get("subject", "")
                reply_subject = f"Re: {reply_subject}" if reply_subject and not reply_subject.startswith("Re:") else reply_subject
                draft = self.gmail.draft_reply(reply["thread_id"], draft_html, reply_subject,
                    to=reply.get("from_addr", ""), sender=sender)
                if draft:
                    self.db.execute("UPDATE replies SET draft_reply_id=? WHERE id=?", (draft.get("id"), reply_id))
            except Exception as draft_err:
                self._log(f"Gmail draft creation failed for reply {reply_id}: {draft_err}")

            self.db.execute("UPDATE replies SET sentiment=?, summary=?, draft_html=? WHERE id=?",
                            (sentiment, summary, draft_html, reply_id))
            self.db.commit()
            return {"success": True, "sentiment": sentiment, "summary": summary, "draft_html": draft_html}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def send_reply_draft(self, reply_id: int, edited_html: str = None) -> dict:
        """Send a reply draft. Uses edited_html if provided, else stored draft_html."""
        reply = self.db.execute("SELECT * FROM replies WHERE id=?", (reply_id,)).fetchone()
        if not reply:
            return {"success": False, "error": "Reply not found"}
        reply = dict(reply)

        body = edited_html if edited_html is not None else reply.get("draft_html", "")
        if not body:
            return {"success": False, "error": "No draft to send"}

        # If the editor gave us plain text, convert newlines to HTML for the send.
        if body and '<' not in body:
            body = body.replace('\n', '<br>')

        subject = reply.get("subject", "")
        to_addr = reply.get("from_addr", "")
        thread_id = reply.get("thread_id", "")

        try:
            reply_subject = f"Re: {subject}" if subject and not subject.startswith("Re:") else subject
            sender = self.default_sender
            sent = self.gmail.send_email(to_addr, reply_subject, body, thread_id=thread_id, sender=sender)
            if sent:
                self.db.mark_reply_handled(reply_id)
                self._log(f"Reply sent to {to_addr}")
                return {"success": True, "message_id": sent.get("id")}
            else:
                return {"success": False, "error": "Failed to send reply"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def blacklist_from_reply(self, reply_id: int) -> dict:
        """Blacklist the sender of a reply."""
        reply = self.db.execute("SELECT from_addr FROM replies WHERE id=?", (reply_id,)).fetchone()
        if not reply:
            return {"success": False, "error": "Reply not found"}
        from_addr = reply[0]
        self.db.blacklist_add(from_addr, "user:reply")
        self.db.mark_reply_handled(reply_id)
        self._log(f"Blacklisted {from_addr} from reply inbox")
        return {"success": True, "email": from_addr}

    # -- Morning Brief --
    def _check_morning_brief(self, now: datetime):
        today_brief = now.replace(hour=MORNING_HOUR, minute=0, second=0, microsecond=0)
        if now < today_brief:
            return
        last = self.db.get_meta("last_morning_brief")
        if last and datetime.fromisoformat(last) >= today_brief:
            return

        brief = self.morning_brief()
        if self.brief_email:
            try:
                self.gmail.send_email(self.brief_email, f"Raj Brief -- {now.strftime('%d %b %Y')}", brief.replace("\n", "<br>"), sender=self.default_sender)
                self._log("Morning brief sent")
            except Exception as e:
                self._log(f"Brief failed: {e}")
        self.db.set_meta("last_morning_brief", now.isoformat())

    def morning_brief(self) -> str:
        today = datetime.now().strftime("%d %b %Y")
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        lines = ["=" * 40, f"RAJ BRIEF -- {today}", "=" * 40, "YESTERDAY"]

        for seq_id in SEQUENCES:
            stats = self.db.execute("SELECT day, COUNT(*) FROM sends WHERE recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?) AND DATE(created_at)=? GROUP BY day", (seq_id, yesterday)).fetchall()
            if stats:
                for day, count in stats: lines.append(f"  {seq_id.upper()} Day {day}: {count} sent")
            else: lines.append(f"  {seq_id.upper()}: No batches")

        replies = self.db.execute("SELECT sentiment, COUNT(*) FROM replies WHERE DATE(received_at)=? OR status IN ('pending','drafted') GROUP BY sentiment", (yesterday,)).fetchall()
        rc = {k: 0 for k in ["positive", "neutral", "hostile", "unsubscribe"]}
        for s, c in replies:
            if s in rc: rc[s] = c
        lines.extend(["", f"REPLIES ({sum(rc.values())} total)", f"  -- {rc['positive']} positive", f"  -- {rc['neutral']} neutral", f"  -- {rc['hostile'] + rc['unsubscribe']} hostile -- blacklisted", "  -> Review drafts in Gmail before sending"])

        bounces = self.db.execute("SELECT email, reason FROM blacklist WHERE DATE(added_at)=? OR reason LIKE 'bounce %'", (yesterday,)).fetchall()
        lines.extend(["", f"BOUNCES ({len(bounces)} overnight)"])
        for email, reason in bounces[:5]: lines.append(f"  {email} -- {reason}")

        lines.extend(["", "DUE TODAY"])
        for seq_id in SEQUENCES:
            for day in SEQUENCES[seq_id]["days"]:
                due = len(self.due_recipients(seq_id, day))
                if due: lines.append(f"  {seq_id.upper()} Day {day}: {due} recipients")

        lines.extend(["", "YOUR ACTIONS", "  1. Review reply drafts in Gmail (DRAFT-ONLY for approval)", "  2. Sequences auto-send at 10 AM -- no action needed", "  3. Reply STOP SCHOOL / STOP CSR / STOP ALL to pause", "=" * 40])
        return "\n".join(lines)

    # -- Emergency Commands --
    def _check_emergency_commands(self, now: datetime):
        last = self.db.get_meta("last_emergency_scan")
        if last and (now - datetime.fromisoformat(last)) < timedelta(minutes=EMERGENCY_INTERVAL): return

        after = int((datetime.now() - timedelta(hours=1)).timestamp())
        msgs = self.gmail.search_messages(f"in:inbox from:me subject:(STOP SCHOOL OR STOP CSR OR STOP ALL OR RESUME) after:{after}", 10)

        for msg in msgs:
            subj = msg.get("subject", "").upper()
            if "STOP SCHOOL" in subj: self.db.set_meta("pause_school", "true"); self._log("SCHOOL paused")
            elif "STOP CSR" in subj: self.db.set_meta("pause_csr", "true"); self._log("CSR paused")
            elif "STOP WSL" in subj: self.db.set_meta("pause_csr-wsl-5", "true"); self._log("CSR-WSL-5 paused")
            elif "STOP ALL" in subj: self.pause(); self._log("ALL paused")
            elif "RESUME" in subj: self.resume(); self.db.execute("DELETE FROM meta WHERE key LIKE 'pause\\_%' ESCAPE '\\'"); self._log("All resumed")

        self.db.set_meta("last_emergency_scan", now.isoformat())

    # -- Campaign State Export --
    def export_campaign_state(self) -> str:
        from pathlib import Path
        now = datetime.now().strftime("%d %b %Y %H:%M")
        lines = [
            f"# Raj Campaign State -- {now}",
            "",
            "## Sequences",
            ""
        ]

        for seq_id in SEQUENCES:
            lines.append(f"### {seq_id.upper()}")
            cfg = SEQUENCES[seq_id]
            for day in cfg["days"]:
                due = self.due_recipients(seq_id, day)
                sent = self.db.execute(
                    "SELECT COUNT(DISTINCT recipient_id) FROM sends WHERE day=? AND status IN ('sent','drafted') AND recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?)",
                    (day, seq_id)
                ).fetchone()[0]
                total = self.db.execute("SELECT COUNT(*) FROM recipients WHERE sequence_id=?", (seq_id,)).fetchone()[0]
                lines.append(f"- Day {day}: {sent}/{total} sent | {len(due)} due")
            lines.append("")

        pending = self.db.execute("SELECT sequence_id, day, COUNT(*) FROM pending_resumes WHERE status='pending' GROUP BY sequence_id, day").fetchall()
        if pending:
            lines.append("## Pending Resumes (Quota Interruptions)")
            for seq_id, day, count in pending:
                lines.append(f"- {seq_id.upper()} Day {day}: {count} emails waiting to resume")
            lines.append("")
        else:
            lines.append("## Pending Resumes")
            lines.append("- None. All batches completed cleanly.")
            lines.append("")

        pending_replies = self.db.execute("SELECT COUNT(*) FROM replies WHERE status='pending'").fetchone()[0]
        drafted_replies = self.db.execute("SELECT COUNT(*) FROM replies WHERE status='drafted'").fetchone()[0]
        lines.append("## Replies")
        lines.append(f"- Pending: {pending_replies}")
        lines.append(f"- Drafted (awaiting approval): {drafted_replies}")
        lines.append("")

        bl_count = self.db.execute("SELECT COUNT(*) FROM blacklist").fetchone()[0]
        bl_recent = self.db.execute("SELECT email, reason FROM blacklist ORDER BY added_at DESC LIMIT 10").fetchall()
        lines.append(f"## Blacklist ({bl_count} total)")
        for email, reason in bl_recent:
            lines.append(f"- `{email}` -- {reason}")
        lines.append("")

        lines.append("## Engine Status")
        lines.append(f"- Running: {self.is_running()}")
        lines.append(f"- Paused: {self.is_paused()}")
        lines.append(f"- Last bounce scan: {self.db.get_meta('last_bounce_scan') or 'Never'}")
        lines.append(f"- Last reply scan: {self.db.get_meta('last_reply_scan') or 'Never'}")
        lines.append("")

        lines.append("---")
        lines.append("*Auto-generated by Raj Campaign Engine*")

        md = "\n".join(lines)

        state_path = Path(__file__).parent / "campaign_state.md"
        with open(state_path, "w", encoding="utf-8") as f:
            f.write(md)

        self._log(f"Campaign state exported to {state_path}")
        return md

    # -- Quota Rollback & Resume --
    def resume_batch(self, seq_id: str, day: int, limit=None) -> BatchResult:
        pending = self.db.execute(
            "SELECT recipient_id, subject FROM pending_resumes WHERE sequence_id=? AND day=? AND status='pending' ORDER BY id",
            (seq_id, day)
        ).fetchall()

        if not pending:
            self._log(f"No pending resumes for {seq_id.upper()} Day {day}")
            return BatchResult(queued=0, sent=0)

        if limit:
            pending = pending[:limit]

        self._log(f"Resuming {seq_id.upper()} Day {day}: {len(pending)} pending")
        sent = 0

        for rec_id, subject in pending:
            rec_row = self.db.execute("SELECT * FROM recipients WHERE id=?", (rec_id,)).fetchone()
            if not rec_row:
                continue
            rec = Recipient(*rec_row)

            subj, body_html, body_text, ab_variant, fmt = self.render(seq_id, day, rec)
            if not subj:
                subj = subject

            try:
                msg = self._send_with_retry(rec.email, subj, body_html, body_text, sender=self.default_sender, format=fmt)
                self.db.campaign_queue_send(rec.id, day, subj, msg.get("id"), "sent", None, ab_variant)
                self.db.commit()
                self.db.execute(
                    "UPDATE pending_resumes SET status='sent', resumed_at=? WHERE recipient_id=? AND sequence_id=? AND day=? AND status='pending'",
                    (datetime.now().isoformat(), rec.id, seq_id, day)
                )
                sent += 1
                self._log(f"Resumed send to {rec.email}")
                time.sleep(SEND_DELAY)
            except Exception as e:
                err = str(e)
                if "quota" in err.lower() or "rate" in err.lower() or "limit" in err.lower():
                    self._log("Rate limit hit again during resume. Stopping.")
                    break
                self._log(f"Resume failed for {rec.email}: {e}")
                self.db.execute(
                    "UPDATE pending_resumes SET status='error', error=? WHERE recipient_id=? AND sequence_id=? AND day=? AND status='pending'",
                    (str(e)[:200], rec.id, seq_id, day)
                )

        self.db.commit()
        self._log(f"Resume complete: {sent}/{len(pending)} sent")
        return BatchResult(queued=len(pending), sent=sent)

    def backdate_sequence(self, seq_id: str, day: int, days_ago: int) -> int:
        cutoff = (datetime.now() - timedelta(days=days_ago)).isoformat()
        rows = self.db.execute(
            "SELECT id, created_at FROM sends WHERE recipient_id IN (SELECT id FROM recipients WHERE sequence_id=?) AND day=? AND created_at > ?",
            (seq_id, day, cutoff)
        ).fetchall()

        count = 0
        for send_id, created_at in rows:
            new_time = (datetime.fromisoformat(created_at) - timedelta(days=days_ago)).isoformat()
            self.db.execute("UPDATE sends SET created_at=? WHERE id=?", (new_time, send_id))
            count += 1

        self.db.commit()
        self._log(f"Backdated {count} sends for {seq_id.upper()} Day {day} by {days_ago} days")
        return count

    def import_blacklist_file(self, filepath: str) -> int:
        from pathlib import Path
        path = Path(filepath)
        if not path.exists():
            self._log(f"Blacklist file not found: {filepath}")
            return 0

        emails = []
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                found = re.findall(r"[\w.+\-]+@[\w.\-]+", line)
                emails.extend(found)

        unique = list(set(e.lower().strip() for e in emails if "@" in e))
        count = 0
        for email in unique:
            if not self.db.blacklist_has(email):
                self.db.blacklist_add(email, f"imported_from_file {path.name}")
                count += 1

        self._log(f"Imported {count} new blacklisted emails from {path.name} ({len(unique)} found)")
        return count
